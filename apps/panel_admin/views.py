from django.views.decorators.http import require_GET, require_http_methods, require_POST
import datetime
import io
import re as _re

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import update_session_auth_hash
from django.db import IntegrityError, transaction
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.db.models import Q

from apps.users.models import Usuario
from apps.ecas.models import Localidad, PuntoECA
from apps.inventory.models import CategoriaMaterial, Material
from apps.panel_admin.service import AdminCatalogService, AdminDashboardService, AdminPuntoECAService
from config import constants as cons


PDF_MIME_TYPE = "application/pdf"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_SUBTITULO_ADMIN = "Panel de Administración — InfoRecicla"
DEFAULT_CITY = "Bogotá"
ADMIN_LISTAR_USUARIOS_URL = "panel_admin:listar_usuarios"
ADMIN_LISTAR_PUBLICACIONES_URL = "panel_admin:listar_publicaciones_admin"
ADMIN_PERFIL_URL = "panel_admin:perfil_admin"
ADMIN_CREATE_PUBLICACION_TEMPLATE = "admin/Publicaciones/createPublicacion.html"
ADMIN_CREATE_USUARIO_TEMPLATE = "admin/Usuarios/createUsuario.html"
ADMIN_EDIT_USUARIO_TEMPLATE = "admin/Usuarios/editUsuario.html"
EXCEL_DESCRIPTION_HEADER = "Descripción"
PUBLICACIONES_NO_HABILITADAS_AJAX_MSG = "El modulo de publicaciones no esta habilitado."
PUBLICACION_NO_ENCONTRADA_MSG = "Publicacion no encontrada."
RECURSO_NO_ENCONTRADO_MSG = "Recurso no encontrado."
CELULAR_ERROR = "El celular debe iniciar con 3 y tener 10 dígitos."
USUARIO_DOCUMENTO_DUPLICADO_MSG = "Ya existe un usuario con ese número de documento."
USUARIO_ACTUALIZADO_OK_MSG = "Usuario actualizado correctamente."
GESTOR_CONTRASENA_REQUERIDA_MSG = "Debe asignar una contrasena al nuevo gestor."
CORREGIR_CAMPOS_MSG = "Corrige los campos señalados."
LISTAR_PUNTOS_ECA_URL = "panel_admin:dashboard_puntos_eca"
LISTAR_CATEGORIAS_PUBLICACION_URL = "panel_admin:listar_categorias_publicacion_admin"
LISTAR_TIPOS_PUBLICACION_URL = "panel_admin:listar_tipos_publicacion_admin"


def _crear_respuesta_descarga(contenido, content_type, filename):
    response = HttpResponse(contenido, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_pdf_context(request):
    import base64
    from django.conf import settings

    logo_path = settings.BASE_DIR / "static" / "img" / "logo.png"
    try:
        logo_b64 = base64.b64encode(logo_path.read_bytes()).decode()
    except Exception:
        logo_b64 = ""

    user = request.user
    if hasattr(user, "nombres") and user.nombres:
        nombre = f"{user.nombres} {user.apellidos}".strip()
    else:
        nombre = user.get_full_name() or user.email or str(user)

    return {"logo_b64": logo_b64, "usuario_generador": nombre}


def _escribir_encabezados_excel(ws, headers):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _validar_nombre_perfil_admin(nombres, errores):
    if not nombres or len(nombres) < 3:
        errores.append("El nombre debe tener al menos 3 caracteres.")
    elif len(nombres) > 30 or not _texto_solo_letras(nombres, permitir_apostrofo=True):
        errores.append("El nombre solo puede contener letras (máx. 30).")


def _validar_apellido_perfil_admin(apellidos, errores):
    if not apellidos or len(apellidos) < 3:
        errores.append("Los apellidos deben tener al menos 3 caracteres.")
    elif len(apellidos) > 40 or not _texto_solo_letras(apellidos, permitir_apostrofo=True):
        errores.append("Los apellidos solo pueden contener letras (máx. 40).")


def _validar_celular_perfil_admin(celular, errores):
    if celular and not _CELULAR.match(celular):
        errores.append(CELULAR_ERROR)


def _validar_ciudad_perfil_admin(ciudad, errores):
    if ciudad and (len(ciudad) > 15 or not _texto_solo_letras(ciudad)):
        errores.append("La ciudad solo puede contener letras (máx. 15).")


def _texto_solo_letras(texto, permitir_apostrofo=False):
    if not texto:
        return False

    caracteres_permitidos = " -'" if permitir_apostrofo else " -"
    return all(caracter.isalpha() or caracter in caracteres_permitidos for caracter in texto)


def _validar_fecha_perfil_admin(fecha_str, errores):
    from datetime import date as date_type
    import re

    if not fecha_str:
        return None

    if re.match(r"^\d{2}-\d{2}-\d{4}$", fecha_str):
        try:
            from datetime import datetime
            fecha_str = datetime.strptime(fecha_str, "%d-%m-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    try:
        hoy = date_type.today()
        fecha_nacimiento = date_type.fromisoformat(fecha_str)
        if fecha_nacimiento > hoy:
            errores.append("La fecha de nacimiento no puede ser futura.")
            return None
        edad = hoy.year - fecha_nacimiento.year - (
            (hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
        )
        if edad < 18:
            errores.append("Debes tener al menos 18 años para registrarte.")
            return None
        if edad > 100:
            errores.append("La fecha debe corresponder a una persona entre 18 y 100 años.")
            return None
        return fecha_nacimiento
    except ValueError:
        errores.append("Formato de fecha inválido.")
        return None


def _validar_localidad_perfil_admin(user, localidad_id, errores):
    localidad_inst = user.localidad
    if localidad_id:
        try:
            localidad_inst = Localidad.objects.get(localidad_id=localidad_id)
        except (Localidad.DoesNotExist, ValueError):
            errores.append("La localidad seleccionada no es válida.")
    return localidad_inst


def _validar_email_perfil_admin(email, user, errores):
    if not email:
        return
    if " " in email:
        errores.append("El correo electrónico no puede contener espacios.")
        return
    if ".." in email:
        errores.append("El correo electrónico no puede contener puntos consecutivos.")
        return
    cantidad_arrobas = email.count("@")
    if cantidad_arrobas != 1:
        errores.append("El correo electrónico debe contener exactamente un símbolo @.")
        return
    dominio = email.rsplit("@", 1)[-1].lower()
    if not dominio.endswith((".com", ".co", ".edu.co", ".com.co", "soy.sena.edu.co", "sena.edu.co")):
        errores.append("El correo electrónico debe terminar en .com, .co, .edu.co, com.co, soy.sena.edu.co o sena.edu.co.")
        return
    if email.lower() != user.email and Usuario.objects.filter(email=email).exists():
        errores.append("Ya existe un usuario con ese correo electrónico.")


def _validar_tipo_documento_perfil_admin(tipo_documento, errores):
    if not tipo_documento:
        return
    if tipo_documento not in {valor for valor, _ in cons.TipoDocumento.choices}:
        errores.append("El tipo de documento seleccionado no es válido.")


def _validar_numero_documento_perfil_admin(numero_documento, user, errores):
    if not numero_documento:
        return
    if not (numero_documento.isdigit() and 6 <= len(numero_documento) <= 20):
        errores.append("El número de documento debe tener entre 6 y 20 dígitos, sin letras ni caracteres especiales.")
    elif numero_documento != user.numero_documento and Usuario.objects.filter(numero_documento=numero_documento).exists():
        errores.append(USUARIO_DOCUMENTO_DUPLICADO_MSG)


def _validar_datos_perfil_admin(user, nombres, apellidos, celular, ciudad, fecha_str, localidad_id,
                                email, tipo_documento, numero_documento):
    errores = []

    _validar_nombre_perfil_admin(nombres, errores)
    _validar_apellido_perfil_admin(apellidos, errores)
    _validar_celular_perfil_admin(celular, errores)
    _validar_ciudad_perfil_admin(ciudad, errores)
    _validar_email_perfil_admin(email, user, errores)
    _validar_tipo_documento_perfil_admin(tipo_documento, errores)
    _validar_numero_documento_perfil_admin(numero_documento, user, errores)
    fecha_nacimiento = _validar_fecha_perfil_admin(fecha_str, errores)
    localidad_inst = _validar_localidad_perfil_admin(user, localidad_id, errores)

    return errores, fecha_nacimiento, localidad_inst


def _validar_cambio_contrasena_admin(user, actual, nueva, confirmar):
    if len(actual) > 128 or len(nueva) > 128 or len(confirmar) > 128:
        return "La contraseña no puede superar los 128 caracteres."
    if not actual or not nueva or not confirmar:
        return "Todos los campos de contraseña son obligatorios."
    if not user.check_password(actual):
        return "La contraseña actual es incorrecta."
    if not _contrasena_cumple_complejidad(nueva):
        return (
            "La nueva contraseña debe tener mínimo 8 caracteres, una mayúscula, "
            "una minúscula, un número y un símbolo (@$!%*?&)."
        )
    if nueva != confirmar:
        return "Las contraseñas nuevas no coinciden."
    return None


def _contrasena_cumple_complejidad(contrasena):
    if len(contrasena) < 8:
        return False

    tiene_mayuscula = any(caracter.isupper() for caracter in contrasena)
    tiene_minuscula = any(caracter.islower() for caracter in contrasena)
    tiene_numero = any(caracter.isdigit() for caracter in contrasena)
    tiene_especial = any(caracter in "@$!%*?&" for caracter in contrasena)

    return tiene_mayuscula and tiene_minuscula and tiene_numero and tiene_especial


def _normalizar_texto(valor, default=""):
    texto = (valor or "").strip()
    return texto if texto else default


def _obtener_datos_usuario_csv(fila):
    email = _normalizar_texto(fila.get("email", "")).lower()
    return {
        "email": email,
        "nombres": _normalizar_texto(fila.get("nombres", "")),
        "apellidos": _normalizar_texto(fila.get("apellidos", "")),
        "celular": _normalizar_texto(fila.get("celular", "")),
        "password": _normalizar_texto(fila.get("password", "")),
        "tipo_usuario": _normalizar_texto(fila.get("tipo_usuario", ""), cons.TipoUsuario.CIUDADANO),
        "tipo_documento": _normalizar_texto(fila.get("tipo_documento", ""), cons.TipoDocumento.CC),
        "numero_documento": _normalizar_texto(fila.get("numero_documento", ""), f"CSV_{email}"),
        "ciudad": _normalizar_texto(fila.get("ciudad", ""), DEFAULT_CITY),
    }


def _validar_datos_usuario_csv(datos, fila_numero):
    errores = []
    if not all([datos["email"], datos["nombres"], datos["apellidos"], datos["celular"], datos["password"]]):
        errores.append(f"Fila {fila_numero}: campos obligatorios incompletos.")
    if datos["email"] and Usuario.objects.filter(email=datos["email"]).exists():
        errores.append(f"Fila {fila_numero}: el email '{datos['email']}' ya existe.")
    if datos["numero_documento"] and Usuario.objects.filter(numero_documento=datos["numero_documento"]).exists():
        errores.append(f"Fila {fila_numero}: el documento '{datos['numero_documento']}' ya existe.")
    return errores


def _crear_usuario_desde_csv(datos):
    with transaction.atomic():
        usuario = Usuario(
            email=datos["email"],
            nombres=datos["nombres"],
            apellidos=datos["apellidos"],
            celular=datos["celular"],
            tipo_usuario=datos["tipo_usuario"],
            tipo_documento=datos["tipo_documento"],
            numero_documento=datos["numero_documento"],
            ciudad=datos["ciudad"],
        )
        usuario.set_password(datos["password"])
        usuario.save()


def _parsear_fecha_crear_usuario(valor):
    if not valor:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(valor, fmt).date()
        except ValueError:
            continue
    return None


def _obtener_datos_crear_usuario_admin(data):
    return {
        "nombres": _normalizar_texto(data.get("nombres", "")),
        "apellidos": _normalizar_texto(data.get("apellidos", "")),
        "email": _normalizar_texto(data.get("email", "")).lower(),
        "celular": _normalizar_texto(data.get("celular", "")),
        "tipo_documento": _normalizar_texto(data.get("tipoDocumento", "")),
        "numero_documento": _normalizar_texto(data.get("numeroDocumento", "")) or None,
        "ciudad": DEFAULT_CITY,
        "localidad_id": _normalizar_texto(data.get("localidad", "")),
        "fecha_nacimiento": _parsear_fecha_crear_usuario(_normalizar_texto(data.get("fechaNacimiento", ""))),
        "tipo_usuario": _normalizar_texto(data.get("tipo_usuario", cons.TipoUsuario.CIUDADANO)),
        "password": data.get("password", ""),
        "password_confirm": data.get("passwordConfirm", ""),
    }


def _validar_texto_personal(valor, campo, min_len, max_len, errores):
    if len(valor) < min_len:
        errores.append(f"El {campo} debe tener al menos {min_len} caracteres.")
    elif len(valor) > max_len:
        errores.append(f"El {campo} no puede superar {max_len} caracteres.")
    elif not _texto_solo_letras(valor, permitir_apostrofo=True):
        errores.append(f"El {campo} solo puede contener letras, espacios, guiones o apóstrofes.")


def _validar_campos_crear_usuario_admin(datos, errores):
    _validar_texto_personal(datos["nombres"], "nombre", 3, 30, errores)
    _validar_texto_personal(datos["apellidos"], "apellido", 3, 40, errores)
    _validar_email_crear_usuario_admin(datos["email"], errores)
    if len(datos["celular"]) != 10 or not datos["celular"].startswith("3"):
        errores.append(CELULAR_ERROR)
    es_admin = datos.get("tipo_usuario") == "ADM"
    if datos["tipo_documento"]:
        if datos["tipo_documento"] not in {valor for valor, _ in cons.TipoDocumento.choices}:
            errores.append("El tipo de documento seleccionado no es válido.")
    elif es_admin:
        errores.append("Debe seleccionar un tipo de documento.")
    if datos["numero_documento"]:
        if not (datos["numero_documento"].isdigit() and 6 <= len(datos["numero_documento"]) <= 20):
            errores.append("El número de documento debe tener entre 6 y 20 dígitos, sin letras ni caracteres especiales.")
    elif es_admin:
        errores.append("Debe ingresar un número de documento.")
    if not datos["ciudad"]:
        errores.append("Debe especificar la ciudad.")


def _validar_email_crear_usuario_admin(email, errores):
    if not email:
        errores.append("Debe ingresar un correo electrónico.")
        return

    if " " in email:
        errores.append("El correo electrónico no puede contener espacios.")
        return

    cantidad_arrobas = email.count("@")
    if cantidad_arrobas != 1:
        errores.append("El correo electrónico debe contener exactamente 1 símbolo @.")
        return

    dominio = email.rsplit("@", 1)[-1].lower()
    if not dominio.endswith((".com", ".co", ".edu.co", ".com.co")):
        errores.append("El correo electrónico debe terminar en .com, .co, .edu.co o .com.co.")


def _validar_credenciales_crear_usuario_admin(datos, errores):
    if not datos["password"] or not datos["password_confirm"]:
        errores.append("Se requiere una contraseña.")
    elif datos["password"] != datos["password_confirm"]:
        errores.append("Las contraseñas no coinciden.")
    elif len(datos["password"]) < 8:
        errores.append("La contraseña debe tener al menos 8 caracteres.")


def _validar_unicidad_crear_usuario_admin(datos, errores):
    if datos["email"] and Usuario.objects.filter(email=datos["email"]).exists():
        errores.append("Ya existe un usuario con ese correo electrónico.")
    if datos["numero_documento"] and Usuario.objects.filter(numero_documento=datos["numero_documento"]).exists():
        errores.append(USUARIO_DOCUMENTO_DUPLICADO_MSG)


def _obtener_localidad_crear_usuario_admin(localidad_id, errores):
    localidad_inst = None
    if localidad_id:
        localidad_inst = Localidad.objects.filter(localidad_id=localidad_id).first()
        if not localidad_inst:
            errores.append("La localidad seleccionada no existe.")
    return localidad_inst


def _validar_tipo_usuario_crear_usuario_admin(datos, errores):
    tipos_validos = {valor for valor, _ in cons.TipoUsuario.choices}
    if datos["tipo_usuario"] not in tipos_validos:
        errores.append("El tipo de usuario seleccionado no es válido.")


def _validar_datos_crear_usuario_admin(datos):
    errores = []

    _validar_campos_crear_usuario_admin(datos, errores)
    _validar_credenciales_crear_usuario_admin(datos, errores)
    _validar_unicidad_crear_usuario_admin(datos, errores)
    localidad_inst = _obtener_localidad_crear_usuario_admin(datos["localidad_id"], errores)
    _validar_tipo_usuario_crear_usuario_admin(datos, errores)

    return errores, localidad_inst


def _crear_usuario_admin_desde_datos(datos, localidad_inst):
    with transaction.atomic():
        usuario = Usuario(
            email=datos["email"],
            numero_documento=datos["numero_documento"],
            nombres=datos["nombres"],
            apellidos=datos["apellidos"],
            celular=datos["celular"],
            tipo_documento=datos["tipo_documento"],
            tipo_usuario=datos["tipo_usuario"],
            ciudad=datos["ciudad"],
            localidad=localidad_inst,
            fecha_nacimiento=datos["fecha_nacimiento"],
        )
        usuario.set_password(datos["password"])
        usuario.save()


def _obtener_datos_crear_punto_eca_admin(data):
    return {
        "nombre_punto": _normalizar_texto(data.get("nombre_punto", "")),
        "nombres": _normalizar_texto(data.get("nombres", "")),
        "apellidos": _normalizar_texto(data.get("apellidos", "")),
        "email": _normalizar_texto(data.get("email", "")).lower(),
        "email_gestor": _normalizar_texto(data.get("email_gestor", "")).lower(),
        "tipo_documento": _normalizar_texto(data.get("tipoDocumento", ""), cons.TipoDocumento.CC),
        "numero_documento": _normalizar_texto(data.get("numeroDocumento", "")),
        "celular": _normalizar_texto(data.get("celular", "")),
        "telefono_punto": _normalizar_texto(data.get("telefono_punto", "")),
        "direccion": _normalizar_texto(data.get("direccion", "")),
        "ciudad": _normalizar_texto(data.get("ciudad", ""), DEFAULT_CITY),
        "localidad_id": _normalizar_texto(data.get("localidad", "")),
        "latitud": _normalizar_texto(data.get("latitud", "")),
        "longitud": _normalizar_texto(data.get("longitud", "")),
        "descripcion": _normalizar_texto(data.get("descripcion", "")),
        "sitio_web": _normalizar_texto(data.get("sitio_web", "")),
        "logo_url_punto": _normalizar_texto(data.get("logo_url_punto", "")),
        "foto_url_punto": _normalizar_texto(data.get("foto_url_punto", "")),
        "horario_atencion": _normalizar_texto(data.get("horario_atencion", "")),
        "password": data.get("password", ""),
        "password_confirm": data.get("passwordConfirm", ""),
    }


def _validar_campos_crear_punto_eca_admin(datos, errores):
    for campo, mensaje in (
        ("nombre_punto", "Debe ingresar el nombre del punto ECA."),
        ("nombres", "Debe ingresar los nombres del gestor."),
        ("apellidos", "Debe ingresar los apellidos del gestor."),
        ("email", "Debe ingresar el email del punto ECA."),
        ("email_gestor", "Debe ingresar el email del gestor."),
        ("direccion", "Debe ingresar la dirección."),
        ("ciudad", "Debe especificar la ciudad."),
    ):
        if not datos[campo]:
            errores.append(mensaje)

    if len(datos["celular"]) != 10 or not datos["celular"].startswith("3"):
        errores.append(CELULAR_ERROR)
    if len(datos["telefono_punto"]) != 10 or not datos["telefono_punto"].startswith("60"):
        errores.append("El teléfono del punto debe iniciar con 60 y tener 10 dígitos.")
    if not datos["latitud"] or not datos["longitud"]:
        errores.append("Debe seleccionar una ubicación en el mapa.")


def _validar_credenciales_crear_punto_eca_admin(datos, errores):
    if not datos["password"] or not datos["password_confirm"]:
        errores.append("Se requiere una contraseña.")
    elif datos["password"] != datos["password_confirm"]:
        errores.append("Las contraseñas no coinciden.")
    elif len(datos["password"]) < 8:
        errores.append("La contraseña debe tener al menos 8 caracteres.")


def _validar_unicidad_crear_punto_eca_admin(datos, errores):
    if datos["email_gestor"] and Usuario.objects.filter(email=datos["email_gestor"]).exists():
        errores.append("Ya existe un usuario con ese correo de gestor.")
    if datos["numero_documento"] and Usuario.objects.filter(numero_documento=datos["numero_documento"]).exists():
        errores.append(USUARIO_DOCUMENTO_DUPLICADO_MSG)


def _obtener_localidad_crear_punto_eca_admin(localidad_id, errores):
    localidad_inst = None
    if localidad_id:
        localidad_inst = Localidad.objects.filter(localidad_id=localidad_id).first()
        if not localidad_inst:
            errores.append("La localidad seleccionada no existe.")
    return localidad_inst


def _validar_datos_crear_punto_eca_admin(datos):
    errores = []

    _validar_campos_crear_punto_eca_admin(datos, errores)
    _validar_credenciales_crear_punto_eca_admin(datos, errores)
    _validar_unicidad_crear_punto_eca_admin(datos, errores)
    localidad_inst = _obtener_localidad_crear_punto_eca_admin(datos["localidad_id"], errores)

    return errores, localidad_inst


def _crear_punto_eca_desde_datos(datos, localidad_inst):
    with transaction.atomic():
        usuario = Usuario(
            email=datos["email_gestor"],
            numero_documento=datos["numero_documento"] or f"GESTORECA_{datos['email_gestor']}",
            celular=datos["celular"],
            nombres=datos["nombres"],
            apellidos=datos["apellidos"],
            tipo_documento=datos["tipo_documento"],
            tipo_usuario=cons.TipoUsuario.GESTOR_ECA,
        )
        usuario.set_password(datos["password"])
        usuario.save()
        PuntoECA.objects.create(
            gestor_eca=usuario,
            nombre=datos["nombre_punto"],
            descripcion=datos["descripcion"],
            telefono_punto=datos["telefono_punto"],
            direccion=datos["direccion"],
            ciudad=datos["ciudad"],
            email=datos["email"],
            celular=datos["celular"],
            logo_url_punto=datos["logo_url_punto"],
            foto_url_punto=datos["foto_url_punto"],
            sitio_web=datos["sitio_web"],
            horario_atencion=datos["horario_atencion"],
            localidad=localidad_inst,
            latitud=float(datos["latitud"]),
            longitud=float(datos["longitud"]),
        )


def _escribir_puntos_eca_excel(ws, puntos):
    for row, punto in enumerate(puntos, 2):
        gestor = f"{punto.gestor_eca.nombres} {punto.gestor_eca.apellidos}" if punto.gestor_eca else ""
        ws.cell(row=row, column=1, value=punto.nombre)
        ws.cell(row=row, column=2, value=punto.direccion or "")
        ws.cell(row=row, column=3, value=punto.localidad.nombre if punto.localidad else "")
        ws.cell(row=row, column=4, value=punto.ciudad or "")
        ws.cell(row=row, column=5, value=punto.telefono_punto or "")
        ws.cell(row=row, column=6, value=punto.email or "")
        ws.cell(row=row, column=7, value=punto.celular or "")
        ws.cell(row=row, column=8, value=gestor)
        ws.cell(row=row, column=9, value=punto.horario_atencion or "")
        ws.cell(row=row, column=10, value=punto.sitio_web or "")
        ws.cell(row=row, column=11, value=punto.estado or "")
        ws.cell(row=row, column=12, value=punto.latitud or "")
        ws.cell(row=row, column=13, value=punto.longitud or "")


def _ajustar_ancho_columnas(ws):
    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)


def _procesar_importar_usuarios_csv(archivo):
    import csv
    import io

    texto = archivo.read().decode("utf-8-sig")
    lector = csv.DictReader(io.StringIO(texto))
    campos_requeridos = {"nombres", "apellidos", "email", "celular", "password"}
    fieldnames = set(lector.fieldnames or [])
    if not campos_requeridos.issubset(fieldnames):
        faltantes = campos_requeridos - fieldnames
        raise ValueError(f"El CSV no tiene las columnas requeridas: {', '.join(faltantes)}")

    creados = 0
    errores = []
    for fila_numero, fila in enumerate(lector, 2):
        datos = _obtener_datos_usuario_csv(fila)
        errores_fila = _validar_datos_usuario_csv(datos, fila_numero)
        if errores_fila:
            errores.extend(errores_fila)
            continue

        try:
            _crear_usuario_desde_csv(datos)
            creados += 1
        except Exception as e:
            errores.append(f"Fila {fila_numero}: {e}")

    return creados, errores


def _procesar_creacion_publicacion_admin(admin_request, categorias, publicaciones_habilitadas):
    from apps.publicaciones.models import CategoriaPublicacion, ImagenPublicacion, Publicacion

    titulo = _normalizar_texto(admin_request.POST.get("titulo"))
    contenido = _normalizar_texto(admin_request.POST.get("contenido"))
    categoria_id = _normalizar_texto(admin_request.POST.get("categoria_id"))

    if not titulo:
        messages.error(admin_request, "El titulo es obligatorio.")
        return None

    resumen = _normalizar_texto(admin_request.POST.get("resumen"))
    if not resumen:
        messages.error(admin_request, "El resumen es obligatorio.")
        return render(
            admin_request,
            ADMIN_CREATE_PUBLICACION_TEMPLATE,
            {
                "publicaciones_habilitadas": publicaciones_habilitadas,
                "categorias": categorias,
                "form_data": admin_request.POST,
                "active_tab": "publicaciones",
            },
        )

    categoria = None
    if categoria_id:
        categoria = CategoriaPublicacion.objects.filter(id=categoria_id).first()
        if not categoria:
            messages.error(admin_request, "La categoria seleccionada no existe.")
            return render(
                admin_request,
                ADMIN_CREATE_PUBLICACION_TEMPLATE,
                {
                    "publicaciones_habilitadas": publicaciones_habilitadas,
                    "categorias": categorias,
                    "form_data": admin_request.POST,
                    "active_tab": "publicaciones",
                },
            )

    limite_bytes = 6 * 1024 * 1024
    imagenes = admin_request.FILES.getlist("imagenes")
    imagenes_grandes = [img.name for img in imagenes if img.size > limite_bytes]
    if imagenes_grandes:
        nombres = ", ".join(imagenes_grandes)
        messages.error(
            admin_request,
            f"Las siguientes imágenes superan el límite de 6 MB y no pueden subirse: {nombres}.",
        )
        return render(
            admin_request,
            ADMIN_CREATE_PUBLICACION_TEMPLATE,
            {
                "publicaciones_habilitadas": publicaciones_habilitadas,
                "categorias": categorias,
                "form_data": admin_request.POST,
            },
        )

    destacado = admin_request.POST.get("destacado") == "1"
    video_url = _normalizar_texto(admin_request.POST.get("video_url"))

    publicacion = Publicacion(
        titulo=titulo,
        contenido=contenido,
        resumen=resumen,
        es_destacado=destacado,
        usuario=admin_request.user,
        categoria=categoria,
        video_url=video_url or None,
        video=admin_request.FILES.get("video") or None,
        video_thumbnail=admin_request.FILES.get("video_thumbnail") or None,
    )
    publicacion.save()

    for imagen in imagenes:
        ImagenPublicacion.objects.create(publicacion=publicacion, imagen=imagen)

    messages.success(admin_request, "Publicacion creada correctamente.")
    return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)


def _procesar_creacion_publicacion_admin_ajax(admin_request):
    from apps.publicaciones.models import CategoriaPublicacion, ImagenPublicacion, Publicacion

    titulo = _normalizar_texto(admin_request.POST.get("titulo"))
    contenido = _normalizar_texto(admin_request.POST.get("contenido"))
    categoria_id = _normalizar_texto(admin_request.POST.get("categoria_id"))
    errores = {}

    if not titulo:
        errores["titulo"] = "El título es obligatorio."

    resumen = _normalizar_texto(admin_request.POST.get("resumen"))
    if not resumen:
        errores["resumen"] = "El resumen es obligatorio."

    categoria = None
    if categoria_id:
        categoria = CategoriaPublicacion.objects.filter(id=categoria_id).first()
        if not categoria:
            errores["categoria_id"] = "La categoría seleccionada no existe."

    limite_bytes = 6 * 1024 * 1024
    imagenes = admin_request.FILES.getlist("imagenes")
    imagenes_grandes = [img.name for img in imagenes if img.size > limite_bytes]
    if imagenes_grandes:
        errores["multimedia"] = f"Las siguientes imágenes superan el límite de 6 MB: {', '.join(imagenes_grandes)}."

    if errores:
        return {"ok": False, "errors": errores, "message": next(iter(errores.values()))}

    destacado = admin_request.POST.get("destacado") == "1"
    video_url = _normalizar_texto(admin_request.POST.get("video_url"))

    try:
        publicacion = Publicacion(
            titulo=titulo,
            contenido=contenido,
            resumen=resumen,
            es_destacado=destacado,
            usuario=admin_request.user,
            categoria=categoria,
            video_url=video_url or None,
            video=admin_request.FILES.get("video") or None,
            video_thumbnail=admin_request.FILES.get("video_thumbnail") or None,
        )
        publicacion.save()

        for imagen in imagenes:
            ImagenPublicacion.objects.create(publicacion=publicacion, imagen=imagen)

        return {"ok": True, "message": "Publicación creada correctamente."}
    except Exception as e:
        return {"ok": False, "errors": {"_general": str(e)}, "message": f"Error al crear la publicación: {e}"}


def _aplicar_datos_usuario_admin(usuario, data):
    usuario.nombres = _normalizar_texto(data.get("nombres"))
    usuario.apellidos = _normalizar_texto(data.get("apellidos"))
    usuario.email = _normalizar_texto(data.get("email")).lower()
    usuario.celular = _normalizar_texto(data.get("celular"))
    ciudad = _normalizar_texto(data.get("ciudad"))
    if ciudad:
        usuario.ciudad = ciudad
    usuario.tipo_usuario = _normalizar_texto(data.get("tipo_usuario"), usuario.tipo_usuario).strip() or usuario.tipo_usuario
    usuario.tipo_documento = _normalizar_texto(data.get("tipoDocumento"), usuario.tipo_documento).strip() or usuario.tipo_documento
    usuario.numero_documento = _normalizar_texto(data.get("numeroDocumento"))
    usuario.biografia = _normalizar_texto(data.get("biografia")) or None

    estado_usuario = _normalizar_texto(data.get("estado_usuario")).lower()
    if estado_usuario in {"activo", "inactivo"}:
        usuario.is_active = estado_usuario == "activo"

    localidad_id = _normalizar_texto(data.get("localidad"))
    if localidad_id:
        usuario.localidad = Localidad.objects.filter(localidad_id=localidad_id).first()

    import re as _re
    fecha_str = _normalizar_texto(data.get("fechaNacimiento"))
    if fecha_str and _re.match(r"^\d{2}-\d{2}-\d{4}$", fecha_str):
        try:
            from datetime import datetime
            fecha_str = datetime.strptime(fecha_str, "%d-%m-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    usuario.fecha_nacimiento = fecha_str or None

    password = _normalizar_texto(data.get("password"))
    password_confirm = _normalizar_texto(data.get("passwordConfirm"))
    if password or password_confirm:
        if password != password_confirm:
            return "Las contrasenas no coinciden."
        usuario.set_password(password)

    return None


def _formatear_error_validacion(excepcion):
    if hasattr(excepcion, "message_dict") and excepcion.message_dict:
        mensajes = []
        for valores in excepcion.message_dict.values():
            if isinstance(valores, (list, tuple)):
                mensajes.extend(str(valor) for valor in valores if str(valor).strip())
            elif str(valores).strip():
                mensajes.append(str(valores))

        if mensajes:
            return " ".join(mensajes)

    if hasattr(excepcion, "messages") and excepcion.messages:
        mensajes = [str(mensaje) for mensaje in excepcion.messages if str(mensaje).strip()]
        if mensajes:
            return " ".join(mensajes)

    return str(excepcion)


def _errores_validacion_lista(excepcion):
    """Returns a LIST of individual error messages instead of one joined string."""
    if hasattr(excepcion, "message_dict") and excepcion.message_dict:
        mensajes = []
        for valores in excepcion.message_dict.values():
            if isinstance(valores, (list, tuple)):
                mensajes.extend(str(valor) for valor in valores if str(valor).strip())
            elif str(valores).strip():
                mensajes.append(str(valores))
        if mensajes:
            return mensajes

    if hasattr(excepcion, "messages") and excepcion.messages:
        mensajes = [str(mensaje) for mensaje in excepcion.messages if str(mensaje).strip()]
        if mensajes:
            return mensajes

    return [str(excepcion)]

def es_administrador(user):
    if not user.is_authenticated:
        return False
    return bool(user.is_staff or user.is_superuser or user.tipo_usuario == cons.TipoUsuario.ADMIN)


@require_http_methods(["GET", "HEAD"])
def admin_redirect_no_autorizado(request):
    return render(request, "base/inicio.html")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def admin(request):
    contexto = {
        "mensaje": "Bienvenido al panel de control de Inforecicla",
        "resumen_general": AdminDashboardService.obtener_resumen_general(),
        "tendencia_usuarios": AdminDashboardService.obtener_tendencia_usuarios(30),
        "dist_puntos_eca": AdminDashboardService.obtener_distribucion_puntos_eca(),
        "dist_materiales": AdminDashboardService.obtener_distribucion_materiales(),
        "alertas_puntos": AdminDashboardService.obtener_alertas_puntos_eca(),
        "puntos_sin_gestor": AdminDashboardService.obtener_puntos_sin_gestor(),
        "gestores_sin_punto": AdminDashboardService.obtener_gestores_sin_punto(),
        "materiales_inactivos": AdminDashboardService.obtener_materiales_inactivos(),
        "publicaciones_pendientes": AdminDashboardService.obtener_publicaciones_pendientes(),
        "materiales_sin_clasificacion": AdminDashboardService.obtener_materiales_sin_clasificacion(),
        "dist_clasificacion_mat": AdminDashboardService.obtener_distribucion_materiales_por_clasificacion(),
    }
    return render(request, "admin/admin.html", contexto)


def usuario_to_dict(usuario):
    return {
        "id": usuario.id,
        "nombres": usuario.nombres,
        "apellidos": usuario.apellidos,
        "email": usuario.email,
        "celular": usuario.celular or "",
        "tipo_usuario": usuario.tipo_usuario,
        "get_tipo_usuario_display": usuario.get_tipo_usuario_display(),
        "is_active": usuario.is_active,
        "ciudad": getattr(usuario, "ciudad", None) or DEFAULT_CITY,
        "localidad_id": str(usuario.localidad.localidad_id)
        if getattr(usuario, "localidad", None)
        else "",
        "tipo_documento": usuario.tipo_documento or "",
        "numero_documento": usuario.numero_documento or "",
        "fecha_nacimiento": usuario.fecha_nacimiento.strftime("%Y-%m-%d")
        if usuario.fecha_nacimiento
        else "",
        "action_url": reverse("panel_admin:editar_usuario_admin", kwargs={"usuario_id": usuario.id}),
    }


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_usuarios(request):
    usuarios = list(Usuario.objects.select_related("localidad").all())
    usuarios_json = [usuario_to_dict(u) for u in usuarios]

    q = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    estado = request.GET.get("estado", "").strip()
    filtered = usuarios
    if q:
        filtered = [u for u in filtered if q.lower() in (u.nombres.lower() + " " + u.apellidos.lower() + " " + u.email.lower())]
    if tipo:
        filtered = [u for u in filtered if u.tipo_usuario == tipo]
    if estado:
        is_active = estado.lower() in ("activo", "true")
        filtered = [u for u in filtered if u.is_active == is_active]

    contexto = {
        "usuarios": filtered,
        "usuarios_json": usuarios_json,
        "search_query": q,
        "search_tipo": tipo,
        "search_estado": estado,
        "localidades": Localidad.objects.all().order_by("nombre"),
        "tipos_documento": cons.TipoDocumento.choices,
        "tipos_usuario": cons.TipoUsuario.choices,
        "dist_roles": AdminDashboardService.obtener_distribucion_usuarios_por_rol(),
        "dist_activos": AdminDashboardService.obtener_distribucion_usuarios_activos(),
        "dist_usuarios_localidad": AdminDashboardService.obtener_distribucion_usuarios_por_localidad(),
        "tendencia_usuarios": AdminDashboardService.obtener_tendencia_usuarios(30),
    }
    return render(request, "admin/Usuarios/listUsuario.html", contexto)


def _filtrar_usuarios_export(request, usuarios):
    q = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    estado = request.GET.get("estado", "").strip()
    if q:
        ql = q.lower()
        usuarios = [u for u in usuarios if ql in (u.nombres.lower() + " " + u.apellidos.lower() + " " + u.email.lower())]
    if tipo:
        usuarios = [u for u in usuarios if u.tipo_usuario == tipo]
    if estado:
        is_active = estado.lower() in ("activo", "true")
        usuarios = [u for u in usuarios if u.is_active == is_active]
    return usuarios


def _filtrar_puntos_eca_export(request, puntos):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        puntos = [p for p in puntos if ql in (p.nombre.lower() + " " + p.direccion.lower() + " " + p.localidad.nombre.lower())]
    localidad = request.GET.get('localidad', '').strip()
    if localidad:
        puntos = [p for p in puntos if p.localidad and p.localidad.nombre == localidad]
    estado = request.GET.get('estado', '').strip()
    if estado:
        puntos = [p for p in puntos if p.estado == estado]
    return puntos


def _filtrar_materiales_export(request, materiales):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        materiales = [m for m in materiales if ql in (m.nombre.lower() + " " + (m.descripcion or "").lower() + " " + (m.categoria.nombre if m.categoria else "").lower())]
    categoria = request.GET.get('categoria', '').strip()
    if categoria:
        materiales = [m for m in materiales if m.categoria and m.categoria.nombre == categoria]
    clasificacion = request.GET.get('clasificacion', '').strip()
    if clasificacion:
        materiales = [m for m in materiales if m.clasificacion and m.clasificacion.upper() == clasificacion.upper()]
    return materiales


def _filtrar_categorias_material_export(request, categorias):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        categorias = [c for c in categorias if ql in (c.nombre.lower() + " " + (c.descripcion or "").lower())]
    estado = request.GET.get('estado', '').strip()
    if estado:
        categorias = [c for c in categorias if c.estado == estado]
    return categorias


def _filtrar_categorias_publicacion_export(request, categorias):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        categorias = [c for c in categorias if ql in (c.nombre.lower() + " " + c.tipo.lower() + " " + (c.descripcion or "").lower())]
    return categorias


def _filtrar_tipos_publicacion_export(request, tipos):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        tipos = [t for t in tipos if ql in (t.nombre.lower() + " " + (t.descripcion or "").lower())]
    estado = request.GET.get('estado', '').strip()
    if estado:
        tipos = [t for t in tipos if t.estado == estado]
    return tipos


def _filtrar_publicaciones_export(request, publicaciones):
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        publicaciones = [p for p in publicaciones if ql in (p.titulo.lower() + " " + (p.contenido or "").lower() + " " + p.usuario.nombres.lower() + " " + p.usuario.apellidos.lower())]
    return publicaciones


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_usuarios_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    usuarios = _filtrar_usuarios_export(request, list(Usuario.objects.all().order_by("apellidos", "nombres")))
    q = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    estado = request.GET.get("estado", "").strip()
    filtros = []
    if q: filtros.append(f"Búsqueda: {q}")
    if tipo: filtros.append(f"Tipo: {tipo}")
    if estado: filtros.append(f"Estado: {estado}")
    ctx = _build_pdf_context(request)
    ctx.update({
        "usuarios": usuarios,
        "titulo_reporte": "Reporte de Usuarios",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Usuarios",
        "total_registros": len(usuarios),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/Usuarios/usuarios_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "usuarios.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_usuarios_excel(request):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["Nombres", "Apellidos", "Email", "Celular", "Tipo Usuario",
               "Tipo Documento", "N° Documento", "Ciudad", "Estado", "Fecha Registro"]
    _escribir_encabezados_excel(ws, headers)

    tipo_labels = dict(cons.TipoUsuario.choices)
    doc_labels = dict(cons.TipoDocumento.choices)

    usuarios = _filtrar_usuarios_export(request, list(Usuario.objects.all().order_by("apellidos", "nombres")))

    for row, u in enumerate(usuarios, 2):
        ws.cell(row=row, column=1, value=u.nombres)
        ws.cell(row=row, column=2, value=u.apellidos)
        ws.cell(row=row, column=3, value=u.email)
        ws.cell(row=row, column=4, value=u.celular or "")
        ws.cell(row=row, column=5, value=str(tipo_labels.get(u.tipo_usuario, u.tipo_usuario)))
        ws.cell(row=row, column=6, value=str(doc_labels.get(u.tipo_documento, u.tipo_documento)))
        ws.cell(row=row, column=7, value=u.numero_documento)
        ws.cell(row=row, column=8, value=u.ciudad or "")
        ws.cell(row=row, column=9, value="Activo" if u.is_active else "Inactivo")
        ws.cell(row=row, column=10, value=u.date_joined.strftime("%Y-%m-%d") if u.date_joined else "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "usuarios.xlsx")


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def importar_usuarios_csv(request):
    if request.method == "GET":
        return redirect(ADMIN_LISTAR_USUARIOS_URL)

    archivo = request.FILES.get("archivo_csv")
    if not archivo:
        messages.error(request, "Debe seleccionar un archivo CSV.")
        return redirect(ADMIN_LISTAR_USUARIOS_URL)
    try:
        creados, errores = _procesar_importar_usuarios_csv(archivo)
    except ValueError as e:
        messages.error(request, f"Error al procesar el archivo: {e}")
        return redirect(ADMIN_LISTAR_USUARIOS_URL)

    if creados:
        messages.success(request, f"{creados} usuario(s) importado(s) correctamente.")
    for err in errores[:10]:
        messages.warning(request, err)
    if len(errores) > 10:
        messages.warning(request, f"... y {len(errores) - 10} error(es) adicionales omitidos.")

    return redirect(ADMIN_LISTAR_USUARIOS_URL)


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_usuario_admin(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    localidades = Localidad.objects.all().order_by("nombre")
    tipos_documento = cons.TipoDocumento.choices
    tipos_usuario = cons.TipoUsuario.choices

    if request.method == "POST":
        data = _obtener_datos_crear_usuario_admin(request.POST)
        errores, localidad_inst = _validar_datos_crear_usuario_admin(data)

        if errores:
            if is_ajax:
                return JsonResponse({"ok": False, "errors": errores, "message": CORREGIR_CAMPOS_MSG})
            return render(
                request,
                ADMIN_CREATE_USUARIO_TEMPLATE,
                {
                    "errores": errores,
                    "localidades": localidades,
                    "tipos_documento": tipos_documento,
                    "tipos_usuario": tipos_usuario,
                    "form_data": request.POST,
                },
            )

        try:
            _crear_usuario_admin_desde_datos(data, localidad_inst)
            if is_ajax:
                return JsonResponse({"ok": True, "message": f"Usuario {data['nombres']} {data['apellidos']} creado correctamente."})
            messages.success(request, f"Usuario {data['nombres']} {data['apellidos']} creado correctamente.")
            return redirect(ADMIN_LISTAR_USUARIOS_URL)
        except (IntegrityError, ValidationError) as e:
            error_msg = f"Error al crear el usuario: {e}"
            if is_ajax:
                return JsonResponse({"ok": False, "errors": [error_msg], "message": error_msg})
            errores = [error_msg]
            return render(
                request,
                ADMIN_CREATE_USUARIO_TEMPLATE,
                {
                    "errores": errores,
                    "localidades": localidades,
                    "tipos_documento": tipos_documento,
                    "tipos_usuario": tipos_usuario,
                    "form_data": request.POST,
                },
            )

    return render(request, ADMIN_CREATE_USUARIO_TEMPLATE, {
        "localidades": localidades,
        "tipos_documento": tipos_documento,
        "tipos_usuario": tipos_usuario,
    })


def publicacion_to_dict(pub):
    return {
        "id": pub.id,
        "titulo": pub.titulo,
        "autor": f"{pub.usuario.nombres} {pub.usuario.apellidos}",
        "categoria": pub.categoria.tipo if pub.categoria else "-",
        "categoria_id": str(pub.categoria.id) if pub.categoria else "",
        "estado": pub.estado,
        "fecha_creacion": pub.fecha_creacion.strftime("%d/%m/%Y") if pub.fecha_creacion else "",
        "contenido": pub.contenido or "",
        "resumen": pub.resumen or "",
        "destacado": pub.es_destacado,
        "video_url": pub.video_url or "",
        "is_active": pub.estado == "ACTIVO",
        "view_url": reverse("panel_admin:ver_publicacion_admin", kwargs={"publicacion_id": pub.id}),
        "action_url": reverse("panel_admin:editar_publicacion_admin", kwargs={"publicacion_id": pub.id}),
    }


def punto_eca_to_dict(punto):
    return {
        "id": punto.id,
        "nombre": punto.nombre,
        "direccion": punto.direccion or "",
        "localidad": punto.localidad.nombre if punto.localidad else "-",
        "localidad_id": str(punto.localidad.localidad_id) if punto.localidad else "",
        "gestor": f"{punto.gestor_eca.nombres} {punto.gestor_eca.apellidos}" if punto.gestor_eca else "Sin gestor",
        "estado": punto.estado,
        "email": punto.email or "",
        "celular": punto.celular or "",
        "telefono_punto": punto.telefono_punto or "",
        "sitio_web": punto.sitio_web or "",
        "horario_atencion": punto.horario_atencion or "",
        "logo_url_punto": punto.logo_url_punto or "",
        "descripcion": punto.descripcion or "",
        "latitud": str(punto.latitud) if punto.latitud else "",
        "longitud": str(punto.longitud) if punto.longitud else "",
        "is_active": punto.estado == "ACTIVO",
        "action_url": reverse("panel_admin:editar_punto_eca_admin", kwargs={"punto_id": punto.id}),
    }


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_publicaciones_admin(request):
    publicaciones = []
    publicaciones_habilitadas = True
    categorias = []
    q = request.GET.get('q', '').strip()
    try:
        from apps.publicaciones.models import CategoriaPublicacion, Publicacion

        all_pubs = list(Publicacion.objects.select_related("usuario", "categoria").all().order_by("-fecha_creacion"))
        categorias = CategoriaPublicacion.objects.all().order_by("nombre", "tipo")
        publicaciones_json = [publicacion_to_dict(p) for p in all_pubs]
        if q:
            ql = q.lower()
            all_pubs = [p for p in all_pubs if ql in (p.titulo.lower() + " " + (p.contenido or "").lower() + " " + p.usuario.nombres.lower() + " " + p.usuario.apellidos.lower())]
        publicaciones = all_pubs
    except Exception:
        publicaciones_habilitadas = False
        publicaciones_json = "[]"

    return render(
        request,
        "admin/Publicaciones/listPublicacion.html",
        {
            "publicaciones": publicaciones,
            "publicaciones_json": publicaciones_json,
            "publicaciones_habilitadas": publicaciones_habilitadas,
            "categorias": categorias,
            "search_query": q,
            "estados": cons.Estado.choices,
            "dist_estado_pub": AdminDashboardService.obtener_distribucion_publicaciones_por_estado(),
            "dist_categoria_pub": AdminDashboardService.obtener_distribucion_publicaciones_por_categoria(),
            "dist_destacadas": AdminDashboardService.obtener_distribucion_publicaciones_destacadas(),
            "tendencia_publicaciones": AdminDashboardService.obtener_tendencia_publicaciones(30),
        },
    )


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_publicaciones_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    try:
        from apps.publicaciones.models import Publicacion
        publicaciones = _filtrar_publicaciones_export(request, list(Publicacion.objects.select_related("usuario", "categoria").all().order_by("-fecha_creacion")))
    except Exception:
        publicaciones = []
    q = request.GET.get("q", "").strip()
    filtros = [f"Búsqueda: {q}"] if q else []
    ctx = _build_pdf_context(request)
    ctx.update({
        "publicaciones": publicaciones,
        "titulo_reporte": "Reporte de Publicaciones",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Publicaciones",
        "total_registros": len(publicaciones),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/Publicaciones/publicaciones_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "publicaciones.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_publicaciones_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publicaciones"

    headers = ["Título", "Resumen", "Categoría", "Autor", "Estado", "Destacado", "Fecha Creación"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    try:
        from apps.publicaciones.models import Publicacion
        publicaciones = _filtrar_publicaciones_export(request, list(Publicacion.objects.select_related("usuario", "categoria").all().order_by("-fecha_creacion")))
    except Exception:
        publicaciones = []

    for row, p in enumerate(publicaciones, 2):
        ws.cell(row=row, column=1, value=p.titulo)
        ws.cell(row=row, column=2, value=p.resumen or "")
        ws.cell(row=row, column=3, value=p.categoria.nombre if p.categoria else "")
        ws.cell(row=row, column=4, value=f"{p.usuario.nombres} {p.usuario.apellidos}" if p.usuario else "")
        ws.cell(row=row, column=5, value=str(p.estado))
        ws.cell(row=row, column=6, value="Sí" if p.es_destacado else "No")
        ws.cell(row=row, column=7, value=p.fecha_creacion.strftime("%Y-%m-%d") if p.fecha_creacion else "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "publicaciones.xlsx")


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_publicacion_admin(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    publicaciones_habilitadas = True
    categorias = []

    try:
        from apps.publicaciones.models import CategoriaPublicacion

        categorias = CategoriaPublicacion.objects.all().order_by("nombre", "tipo")

        if request.method == "POST":
            if is_ajax:
                resultado = _procesar_creacion_publicacion_admin_ajax(request)
                return JsonResponse(resultado)
            respuesta = _procesar_creacion_publicacion_admin(request, categorias, publicaciones_habilitadas)
            if respuesta is not None:
                return respuesta

    except Exception as e:
        publicaciones_habilitadas = False
        if is_ajax:
            return JsonResponse({"ok": False, "errors": [f"No se pudo cargar el modulo de publicaciones: {e}"], "message": str(e)})
        messages.error(request, f"No se pudo cargar el modulo de publicaciones: {e}")

    return render(
        request,
        ADMIN_CREATE_PUBLICACION_TEMPLATE,
        {
            "publicaciones_habilitadas": publicaciones_habilitadas,
            "categorias": categorias,
            "active_tab": "publicaciones",
        },
    )


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_puntos_eca_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    puntos = _filtrar_puntos_eca_export(request, list(PuntoECA.objects.select_related("gestor_eca", "localidad").all().order_by("nombre")))
    q = request.GET.get("q", "").strip()
    localidad = request.GET.get("localidad", "").strip()
    estado = request.GET.get("estado", "").strip()
    filtros = []
    if q: filtros.append(f"Búsqueda: {q}")
    if localidad: filtros.append(f"Localidad: {localidad}")
    if estado: filtros.append(f"Estado: {estado}")
    ctx = _build_pdf_context(request)
    ctx.update({
        "puntos": puntos,
        "titulo_reporte": "Reporte de Puntos ECA",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Puntos ECA",
        "total_registros": len(puntos),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/PuntoECA/puntos_eca_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "puntos_eca.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_puntos_eca_excel(request):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Puntos ECA"

    headers = ["Nombre", "Dirección", "Localidad", "Ciudad", "Teléfono", "Email",
               "Celular", "Gestor", "Horario", "Sitio Web", "Estado", "Latitud", "Longitud"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    puntos = _filtrar_puntos_eca_export(request, list(PuntoECA.objects.select_related("gestor_eca", "localidad").all().order_by("nombre")))
    _escribir_puntos_eca_excel(ws, puntos)
    _ajustar_ancho_columnas(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "puntos_eca.xlsx")


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_punto_eca_admin(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    localidades = Localidad.objects.all().order_by("nombre")
    tipos_documento = cons.TipoDocumento.choices

    if request.method == "POST":
        data = _obtener_datos_crear_punto_eca_admin(request.POST)
        errores, localidad_inst = _validar_datos_crear_punto_eca_admin(data)

        if errores:
            if is_ajax:
                return JsonResponse({"ok": False, "errors": errores, "message": CORREGIR_CAMPOS_MSG})
            return render(
                request,
                "admin/PuntoECA/createPuntoECA.html",
                {
                    "errores": errores,
                    "localidades": localidades,
                    "tipos_documento": tipos_documento,
                    "form_data": request.POST,
                },
            )

        try:
            _crear_punto_eca_desde_datos(data, localidad_inst)
            if is_ajax:
                return JsonResponse({"ok": True, "message": f"Punto ECA '{data['nombre_punto']}' creado correctamente."})
            messages.success(request, f"Punto ECA '{data['nombre_punto']}' creado correctamente.")
            return redirect(LISTAR_PUNTOS_ECA_URL)
        except (IntegrityError, ValidationError) as e:
            error_msg = f"Error al crear el punto ECA: {e}"
            if is_ajax:
                return JsonResponse({"ok": False, "errors": [error_msg], "message": error_msg})
            messages.error(request, error_msg)

    return render(request, "admin/PuntoECA/createPuntoECA.html", {
        "localidades": localidades,
        "tipos_documento": tipos_documento,
    })


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_puntos_eca_admin(request):
    all_puntos = list(PuntoECA.objects.select_related("gestor_eca", "localidad").all().order_by("nombre"))
    puntos_json = [punto_eca_to_dict(p) for p in all_puntos]
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        all_puntos = [p for p in all_puntos if ql in (p.nombre.lower() + " " + p.direccion.lower() + " " + p.localidad.nombre.lower())]
    return render(request, "admin/PuntoECA/listPuntoECA.html", {
        "puntos": all_puntos,
        "puntos_json": puntos_json,
        "search_query": q,
        "localidades": Localidad.objects.all().order_by("nombre"),
        "tipos_documento": cons.TipoDocumento.choices,
        "estados": cons.Estado.choices,
        "dist_estado_punto": AdminDashboardService.obtener_distribucion_puntos_eca_por_estado(),
        "dist_gestor_punto": AdminDashboardService.obtener_distribucion_puntos_eca_con_gestor(),
    })


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def puntos_eca_dashboard(request):
    contexto = {
        "puntos_dashboard": AdminPuntoECAService.obtener_puntos_dashboard(),
        "historial": AdminPuntoECAService.obtener_historial(),
        "eventos": AdminPuntoECAService.obtener_eventos(),
        "conversaciones": AdminPuntoECAService.obtener_conversaciones(),
        "usuarios": AdminPuntoECAService.obtener_usuarios_admin(),
        "kpis": AdminPuntoECAService.obtener_kpis(),
        "inv_data": AdminPuntoECAService.obtener_inventario_desglosado(),
        "localidades": Localidad.objects.all().order_by("nombre"),
        "estados": cons.Estado.choices,
    }
    return render(request, "admin/PuntoECA/dashboard.html", contexto)


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_materiales_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    materiales = _filtrar_materiales_export(request, list(Material.objects.select_related("categoria").all().order_by("nombre")))
    q = request.GET.get("q", "").strip()
    categoria = request.GET.get("categoria", "").strip()
    clasificacion = request.GET.get("clasificacion", "").strip()
    filtros = []
    if q: filtros.append(f"Búsqueda: {q}")
    if categoria: filtros.append(f"Categoría: {categoria}")
    if clasificacion: filtros.append(f"Clasificación: {clasificacion}")
    ctx = _build_pdf_context(request)
    ctx.update({
        "materiales": materiales,
        "titulo_reporte": "Reporte de Materiales Reciclables",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Materiales",
        "total_registros": len(materiales),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/Materiales/materiales_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "materiales.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_materiales_excel(request):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales"

    headers = ["Nombre", EXCEL_DESCRIPTION_HEADER, "Categoría", "Clasificación", "Estado"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    materiales = _filtrar_materiales_export(request, list(Material.objects.select_related("categoria").all().order_by("nombre")))
    for row, m in enumerate(materiales, 2):
        ws.cell(row=row, column=1, value=m.nombre)
        ws.cell(row=row, column=2, value=m.descripcion or "")
        ws.cell(row=row, column=3, value=m.categoria.nombre if m.categoria else "")
        ws.cell(row=row, column=4, value=m.clasificacion or "")
        ws.cell(row=row, column=5, value=m.estado or "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "materiales.xlsx")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_materiales_admin(request):
    all_materiales = list(Material.objects.select_related("categoria").all().order_by("nombre"))
    materiales_json = [material_to_dict(m) for m in all_materiales]
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        all_materiales = [m for m in all_materiales if ql in (m.nombre.lower() + " " + (m.descripcion or "").lower() + " " + (m.categoria.nombre if m.categoria else "").lower())]
    return render(request, "admin/Materiales/listMaterial.html", {"materiales": all_materiales, "materiales_json": materiales_json, "search_query": q})


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_categorias_material_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    categorias = _filtrar_categorias_material_export(request, list(CategoriaMaterial.objects.all().order_by("nombre")))
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip()
    filtros = []
    if q: filtros.append(f"Búsqueda: {q}")
    if estado: filtros.append(f"Estado: {estado}")
    ctx = _build_pdf_context(request)
    ctx.update({
        "categorias": categorias,
        "titulo_reporte": "Reporte de Categorías de Materiales",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Categorías de Materiales",
        "total_registros": len(categorias),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/CategoriasMateriales/categorias_material_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "categorias_material.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_categorias_material_excel(request):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías de Materiales"

    headers = ["Nombre", EXCEL_DESCRIPTION_HEADER, "Estado"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    categorias = _filtrar_categorias_material_export(request, list(CategoriaMaterial.objects.all().order_by("nombre")))
    for row, c in enumerate(categorias, 2):
        ws.cell(row=row, column=1, value=c.nombre)
        ws.cell(row=row, column=2, value=c.descripcion or "")
        ws.cell(row=row, column=3, value=c.estado or "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "categorias_material.xlsx")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_categorias_material_admin(request):
    all_categorias = list(CategoriaMaterial.objects.all().order_by("nombre"))
    categorias_json = [categoria_material_to_dict(c) for c in all_categorias]
    q = request.GET.get('q', '').strip()
    if q:
        ql = q.lower()
        all_categorias = [c for c in all_categorias if ql in (c.nombre.lower() + " " + (c.descripcion or "").lower())]
    return render(request, "admin/CategoriasMateriales/listCategoriaMaterial.html", {"categorias": all_categorias, "categorias_json": categorias_json, "search_query": q})


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_categorias_publicacion_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    try:
        from apps.publicaciones.models import CategoriaPublicacion
        categorias = _filtrar_categorias_publicacion_export(request, list(CategoriaPublicacion.objects.all().order_by("tipo")))
    except Exception:
        categorias = []
    q = request.GET.get("q", "").strip()
    filtros = [f"Búsqueda: {q}"] if q else []
    ctx = _build_pdf_context(request)
    ctx.update({
        "categorias": categorias,
        "titulo_reporte": "Reporte de Categorías de Publicaciones",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Categorías de Publicaciones",
        "total_registros": len(categorias),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/CategoriasPublicaciones/categorias_publicacion_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "categorias_publicacion.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_categorias_publicacion_excel(request):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías de Publicaciones"

    headers = ["Tipo", EXCEL_DESCRIPTION_HEADER, "Estado"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    try:
        from apps.publicaciones.models import CategoriaPublicacion
        categorias = _filtrar_categorias_publicacion_export(request, list(CategoriaPublicacion.objects.all().order_by("tipo")))
    except Exception:
        categorias = []

    for row, c in enumerate(categorias, 2):
        ws.cell(row=row, column=1, value=c.tipo)
        ws.cell(row=row, column=2, value=c.descripcion or "")
        ws.cell(row=row, column=3, value=c.estado or "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "categorias_publicacion.xlsx")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_categorias_publicacion_admin(request):
    categorias = []
    categorias_json = "[]"
    publicaciones_habilitadas = True
    q = request.GET.get('q', '').strip()
    try:
        from apps.publicaciones.models import CategoriaPublicacion

        all_categorias = list(CategoriaPublicacion.objects.all().order_by("nombre", "tipo"))
        categorias_json = [categoria_publicacion_to_dict(c) for c in all_categorias]
        if q:
            ql = q.lower()
            all_categorias = [c for c in all_categorias if ql in (c.nombre.lower() + " " + c.tipo.lower() + " " + (c.descripcion or "").lower())]
        categorias = all_categorias
    except Exception:
        publicaciones_habilitadas = False

    return render(
        request,
        "admin/CategoriasPublicaciones/listCategoriaPublicacion.html",
        {
            "categorias": categorias,
            "categorias_json": categorias_json,
            "publicaciones_habilitadas": publicaciones_habilitadas,
            "search_query": q,
            "active_tab": "categorias_publicacion",
            "tipos_publicacion": AdminCatalogService._tipos_publicacion_disponibles(),
            "estados": cons.Estado.choices,
            "dist_tipo_cat_pub": AdminDashboardService.obtener_distribucion_categorias_publicacion_por_tipo(),
            "dist_estado_cat_pub": AdminDashboardService.obtener_distribucion_categorias_publicacion_por_estado(),
            "tendencia_publicaciones": AdminDashboardService.obtener_tendencia_publicaciones(30),
        },
    )


def _contexto_usuario_admin(usuario):
    return {
        "usuario": usuario,
        "localidades": Localidad.objects.all().order_by("nombre"),
        "tipos_documento": cons.TipoDocumento.choices,
        "tipos_usuario": cons.TipoUsuario.choices,
    }


def _procesar_post_usuario_admin(request, usuario, is_ajax):
    error = _aplicar_datos_usuario_admin(usuario, request.POST)
    if error:
        return _respuesta_error_usuario(request, is_ajax, error)

    try:
        usuario.full_clean()
        usuario.save()
        return _respuesta_exito_usuario(request, is_ajax, usuario)
    except ValidationError as e:
        return _manejar_error_validacion(request, is_ajax, e)
    except Exception as e:
        if is_ajax:
            return JsonResponse({"ok": False, "errors": [str(e)], "message": str(e)})
        messages.error(request, f"No se pudo actualizar el usuario: {e}")
        return None


def _respuesta_error_usuario(request, is_ajax, error):
    if is_ajax:
        return JsonResponse({"ok": False, "errors": [error], "message": error})
    messages.error(request, error)
    return None


def _respuesta_exito_usuario(request, is_ajax, usuario):
    if is_ajax:
        return JsonResponse({"ok": True, "message": USUARIO_ACTUALIZADO_OK_MSG})
    messages.success(request, USUARIO_ACTUALIZADO_OK_MSG)
    return redirect("panel_admin:editar_usuario_admin", usuario_id=usuario.id)


def _manejar_error_validacion(request, is_ajax, excepcion):
    lista_errores = _errores_validacion_lista(excepcion)
    if is_ajax:
        return JsonResponse({"ok": False, "errors": lista_errores, "message": CORREGIR_CAMPOS_MSG})
    mensajes_error = " ".join(lista_errores)
    messages.error(request, f"No se pudo actualizar el usuario: {mensajes_error}")
    return None


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_usuario_admin(request, usuario_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    usuario = Usuario.objects.filter(id=usuario_id).first()
    if not usuario:
        if is_ajax:
            return JsonResponse({"ok": False, "message": "Usuario no encontrado."})
        messages.error(request, "Usuario no encontrado.")
        return redirect(ADMIN_LISTAR_USUARIOS_URL)

    if request.method == "POST":
        respuesta = _procesar_post_usuario_admin(request, usuario, is_ajax)
        if respuesta:
            return respuesta

    return render(request, ADMIN_EDIT_USUARIO_TEMPLATE, _contexto_usuario_admin(usuario))


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_publicacion_admin(request, publicacion_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    try:
        from apps.publicaciones.models import CategoriaPublicacion, Publicacion
    except Exception:
        if is_ajax:
            return JsonResponse({"ok": False, "message": PUBLICACIONES_NO_HABILITADAS_AJAX_MSG})
        messages.error(request, "El modulo de publicaciones no esta habilitado en la configuracion actual.")
        return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)

    publicacion = Publicacion.objects.select_related("categoria", "usuario").filter(id=publicacion_id).first()
    if not publicacion:
        if is_ajax:
            return JsonResponse({"ok": False, "message": PUBLICACION_NO_ENCONTRADA_MSG})
        messages.error(request, PUBLICACION_NO_ENCONTRADA_MSG)
        return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)

    if request.method == "POST":
        resultado = AdminCatalogService.actualizar_publicacion(publicacion_id, request.POST, request.FILES)
        if is_ajax:
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
            return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)
        messages.error(request, resultado["message"])
        publicacion.refresh_from_db()

    categorias = CategoriaPublicacion.objects.all().order_by("tipo")
    return render(
        request,
        "admin/Publicaciones/editPublicacion.html",
        {
            "publicacion": publicacion,
            "categorias": categorias,
            "estados": cons.Estado.choices,
        },
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET"])
def ver_publicacion_admin(request, publicacion_id):
    try:
        from apps.publicaciones.models import Publicacion
    except Exception:
        messages.error(request, PUBLICACIONES_NO_HABILITADAS_AJAX_MSG)
        return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)

    publicacion = (
        Publicacion.objects
        .select_related("categoria", "usuario")
        .prefetch_related("imagenes", "comentarios", "reacciones")
        .filter(id=publicacion_id)
        .first()
    )
    if not publicacion:
        messages.error(request, PUBLICACION_NO_ENCONTRADA_MSG)
        return redirect(ADMIN_LISTAR_PUBLICACIONES_URL)

    return render(
        request,
        "admin/Publicaciones/showPublicacion.html",
        {"publicacion": publicacion, "active_tab": "publicaciones"},
    )


def _actualizar_gestor_eca(punto, request):
    data = {k: (request.POST.get(k) or "").strip() for k in (
        "nombres", "apellidos", "email_gestor", "tipoDocumento",
        "numeroDocumento", "celular", "password",
    )}
    nombres, apellidos, email, tipo_doc, num_doc, celular, password = (
        data["nombres"], data["apellidos"], data["email_gestor"],
        data["tipoDocumento"], data["numeroDocumento"], data["celular"], data["password"],
    )
    if not (nombres or apellidos or email):
        return None

    gestor = punto.gestor_eca
    if gestor:
        gestor.nombres = nombres or gestor.nombres
        gestor.apellidos = apellidos or gestor.apellidos
        gestor.tipo_documento = tipo_doc or gestor.tipo_documento
        gestor.numero_documento = num_doc or gestor.numero_documento
        gestor.celular = celular or gestor.celular
        gestor.save()
        return None

    if not password:
        return GESTOR_CONTRASENA_REQUERIDA_MSG

    gestor = Usuario(
        nombres=nombres, apellidos=apellidos,
        email=email or f"gestor_{punto.id}@eca.com",
        numero_documento=num_doc or f"GESTORECA_{punto.id}",
        tipo_documento=tipo_doc or cons.TipoDocumento.CC,
        tipo_usuario=cons.TipoUsuario.GESTOR_ECA,
        celular=celular,
    )
    gestor.set_password(password)
    gestor.save()
    punto.gestor_eca = gestor
    punto.save(update_fields=["gestor_eca"])
    return None


def _procesar_edicion_punto(punto, punto_id, request, is_ajax):
    resultado = None
    error_gestor = None
    try:
        with transaction.atomic():
            error_gestor = _actualizar_gestor_eca(punto, request)
            if not error_gestor:
                resultado = AdminCatalogService.actualizar_punto_eca(punto_id, request.POST)
    except (IntegrityError, ValidationError) as e:
        error_gestor = f"Error al actualizar: {e}"

    if error_gestor:
        punto.refresh_from_db()
        if is_ajax:
            return JsonResponse({"ok": False, "message": error_gestor})
        messages.error(request, error_gestor)
        return render(
            request, "admin/PuntoECA/editPuntoECA.html",
            {"punto": punto, "localidades": Localidad.objects.all().order_by("nombre"),
             "estados": cons.Estado.choices, "tipos_documento": cons.TipoDocumento.choices,
             "form_data": request.POST, "errores": [error_gestor]},
        )

    if resultado is None:
        return None

    if is_ajax:
        return JsonResponse(resultado)
    if resultado["ok"]:
        messages.success(request, resultado["message"])
        return redirect(LISTAR_PUNTOS_ECA_URL)
    messages.error(request, resultado["message"])
    punto.refresh_from_db()
    return None


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_punto_eca_admin(request, punto_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    punto = PuntoECA.objects.select_related("localidad", "gestor_eca").filter(id=punto_id).first()
    if not punto:
        if is_ajax:
            return JsonResponse({"ok": False, "message": "Punto ECA no encontrado."})
        messages.error(request, "Punto ECA no encontrado.")
        return redirect(LISTAR_PUNTOS_ECA_URL)

    if request.method == "POST":
        respuesta = _procesar_edicion_punto(punto, punto_id, request, is_ajax)
        if respuesta is not None:
            return respuesta

    return render(
        request,
        "admin/PuntoECA/editPuntoECA.html",
        {
            "punto": punto,
            "localidades": Localidad.objects.all().order_by("nombre"),
            "estados": cons.Estado.choices,
            "tipos_documento": cons.TipoDocumento.choices,
        },
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_material_admin(request, material_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    material = Material.objects.select_related("categoria").filter(id=material_id).first()
    if not material:
        if is_ajax:
            return JsonResponse({"ok": False, "message": "Material no encontrado."})
        messages.error(request, "Material no encontrado.")
        return redirect("panel_admin:listar_materiales_admin")

    if request.method == "POST":
        resultado = AdminCatalogService.actualizar_material(material_id, request.POST, request.FILES)
        if is_ajax:
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
            return redirect("/panel_admin/materiales/gestion/?tab=materiales")
        messages.error(request, resultado["message"])
        material.refresh_from_db()

    return render(
        request,
        "admin/Materiales/editMaterial.html",
        {
            "material": material,
            "categorias": CategoriaMaterial.objects.all().order_by("nombre"),
            "estados": cons.Estado.choices,
        },
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_categoria_material_admin(request, categoria_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    categoria = CategoriaMaterial.objects.filter(id=categoria_id).first()
    if not categoria:
        if is_ajax:
            return JsonResponse({"ok": False, "message": "Categoria de material no encontrada."})
        messages.error(request, "Categoria de material no encontrada.")
        return redirect("panel_admin:listar_categorias_material_admin")

    if request.method == "POST":
        resultado = AdminCatalogService.actualizar_categoria_material(categoria_id, request.POST)
        if is_ajax:
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
            return redirect("/panel_admin/materiales/gestion/?tab=categorias")
        messages.error(request, resultado["message"])
        categoria.refresh_from_db()

    return render(
        request,
        "admin/CategoriasMateriales/editCategoriaMaterial.html",
        {
            "categoria": categoria,
            "estados": cons.Estado.choices,
        },
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_categoria_publicacion_admin(request, categoria_id):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    try:
        from apps.publicaciones.models import CategoriaPublicacion
    except Exception:
        if is_ajax:
            return JsonResponse({"ok": False, "message": PUBLICACIONES_NO_HABILITADAS_AJAX_MSG})
        messages.error(request, "El modulo de publicaciones no esta habilitado en la configuracion actual.")
        return redirect(LISTAR_CATEGORIAS_PUBLICACION_URL)

    categoria = CategoriaPublicacion.objects.filter(id=categoria_id).first()
    if not categoria:
        if is_ajax:
            return JsonResponse({"ok": False, "message": "Categoria de publicacion no encontrada."})
        messages.error(request, "Categoria de publicacion no encontrada.")
        return redirect(LISTAR_CATEGORIAS_PUBLICACION_URL)

    tipos_publicacion = AdminCatalogService._tipos_publicacion_disponibles()
    form_data = {
        "nombre": getattr(categoria, "nombre", ""),
        "descripcion": getattr(categoria, "descripcion", ""),
        "tipo": categoria.tipo if categoria.tipo in {v for v, _ in tipos_publicacion} else "",
        "estado": categoria.estado,
    }

    if request.method == "POST":
        resultado = AdminCatalogService.actualizar_categoria_publicacion(categoria_id, request.POST)
        if is_ajax:
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
            return redirect(LISTAR_CATEGORIAS_PUBLICACION_URL)
        messages.error(request, resultado["message"])
        form_data = {
            "nombre": request.POST.get("nombre", ""),
            "descripcion": request.POST.get("descripcion", ""),
            "tipo": request.POST.get("tipo", ""),
            "estado": request.POST.get("estado", ""),
        }

    return render(
        request,
        "admin/CategoriasPublicaciones/editCategoriaPublicacion.html",
        {
            "categoria": categoria,
            "form_data": form_data,
            "tipos_publicacion": tipos_publicacion,
            "estados": cons.Estado.choices,
            "active_tab": "categorias_publicacion",
        },
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_categoria_material(request):
    if request.method == "POST":
        resultado = AdminCatalogService.crear_categoria_material(request.POST)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
        else:
            messages.error(request, resultado["message"])
        return redirect("/panel_admin/materiales/gestion/?tab=categorias")

    return render(
        request,
        "admin/CategoriasMateriales/createCategoriaMaterial.html",
        {"estados": cons.Estado.choices},
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_material_admin(request):
    if request.method == "POST":
        resultado = AdminCatalogService.crear_material(request.POST, request.FILES)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
        else:
            messages.error(request, resultado["message"])
        return redirect("/panel_admin/materiales/gestion/?tab=materiales")

    return render(
        request,
        "admin/Materiales/createMaterial.html",
        {
            "estados": cons.Estado.choices,
            "todas_categorias": CategoriaMaterial.objects.all().order_by("nombre"),
        },
    )


def material_to_dict(m):
    return {
        "id": m.id,
        "nombre": m.nombre,
        "descripcion": m.descripcion or "",
        "categoria_nombre": m.categoria.nombre if m.categoria else "-",
        "categoria_id": str(m.categoria.id) if m.categoria else "",
        "clasificacion": m.clasificacion,
        "estado": m.estado,
        "is_active": m.estado == "ACTIVO",
        "action_url": reverse("panel_admin:editar_material_admin", kwargs={"material_id": m.id}),
    }


def categoria_material_to_dict(c):
    return {
        "id": c.id,
        "nombre": c.nombre,
        "descripcion": c.descripcion or "",
        "estado": c.estado,
    }


def categoria_publicacion_to_dict(c):
    return {
        "id": c.id,
        "nombre": c.nombre or c.tipo,
        "tipo": c.tipo,
        "get_tipo_display": c.get_tipo_display(),
        "descripcion": c.descripcion or "",
        "estado": c.estado,
        "is_active": c.estado == "ACTIVO",
        "action_url": reverse("panel_admin:editar_categoria_publicacion_admin", kwargs={"categoria_id": c.id}),
    }


def tipo_publicacion_to_dict(t):
    return {
        "id": t.id,
        "nombre": t.nombre,
        "descripcion": t.descripcion or "",
        "estado": t.estado,
        "is_active": t.estado == "ACTIVO",
        "action_url": reverse("panel_admin:editar_tipo_publicacion_admin", kwargs={"tipo_id": t.id}),
    }


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def listar_tipos_publicacion_admin(request):
    try:
        from apps.publicaciones.models import TipoPublicacion

        all_tipos = _filtrar_tipos_publicacion_export(request, list(TipoPublicacion.objects.all().order_by("nombre")))
        tipos_json = [tipo_publicacion_to_dict(t) for t in all_tipos]
    except Exception:
        all_tipos = []
        tipos_json = []
    from collections import Counter
    dist_estado = Counter(t.estado for t in all_tipos)
    dist_estado_tipo_pub = [{"estado": k, "count": v} for k, v in dist_estado.items()]
    return render(
        request,
        "admin/Publicaciones/listTipoPublicacion.html",
        {
            "tipos": all_tipos,
            "tipos_json": tipos_json,
            "active_tab": "tipos_publicacion",
            "estados": cons.Estado.choices,
            "dist_estado_tipo_pub": dist_estado_tipo_pub,
            "search_query": request.GET.get("q", "").strip(),
        },
    )


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_tipos_publicacion_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML

    try:
        from apps.publicaciones.models import TipoPublicacion

        tipos = _filtrar_tipos_publicacion_export(request, list(TipoPublicacion.objects.all().order_by("nombre")))
    except Exception:
        tipos = []
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip()
    filtros = []
    if q: filtros.append(f"Búsqueda: {q}")
    if estado: filtros.append(f"Estado: {estado}")
    ctx = _build_pdf_context(request)
    ctx.update({
        "tipos": tipos,
        "titulo_reporte": "Reporte de Tipos de Publicación",
        "subtitulo_reporte": PDF_SUBTITULO_ADMIN,
        "tipo_reporte": "Tipos de Publicación",
        "total_registros": len(tipos),
        "filtros_activos": " | ".join(filtros) if filtros else "Ninguno",
    })
    html = render_to_string("admin/Publicaciones/tipos_publicacion_pdf.html", ctx)
    pdf = HTML(string=html).write_pdf()
    return _crear_respuesta_descarga(pdf, PDF_MIME_TYPE, "tipos_publicacion.pdf")


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def exportar_tipos_publicacion_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tipos de Publicación"

    headers = ["Nombre", "Descripción", "Estado"]
    fill = PatternFill(start_color="1A7A3A", end_color="1A7A3A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    try:
        from apps.publicaciones.models import TipoPublicacion

        tipos = _filtrar_tipos_publicacion_export(request, list(TipoPublicacion.objects.all().order_by("nombre")))
    except Exception:
        tipos = []

    for row, t in enumerate(tipos, 2):
        ws.cell(row=row, column=1, value=t.nombre)
        ws.cell(row=row, column=2, value=t.descripcion or "")
        ws.cell(row=row, column=3, value=t.estado or "")

    for col in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _crear_respuesta_descarga(buf.read(), XLSX_MIME_TYPE, "tipos_publicacion.xlsx")


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_tipo_publicacion(request):
    if request.method == "POST":
        resultado = AdminCatalogService.crear_tipo_publicacion(request.POST)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
        else:
            messages.error(request, resultado["message"])
        return redirect(LISTAR_TIPOS_PUBLICACION_URL)
    return render(
        request,
        "admin/Publicaciones/createTipoPublicacion.html",
        {"estados": cons.Estado.choices},
    )


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def editar_tipo_publicacion_admin(request, tipo_id):
    try:
        from apps.publicaciones.models import TipoPublicacion

        tipo = TipoPublicacion.objects.filter(id=tipo_id).first()
    except Exception:
        tipo = None
    if not tipo:
        messages.error(request, RECURSO_NO_ENCONTRADO_MSG)
        return         redirect(LISTAR_TIPOS_PUBLICACION_URL)

    if request.method == "POST":
        resultado = AdminCatalogService.actualizar_tipo_publicacion(tipo_id, request.POST)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
        else:
            messages.error(request, resultado["message"])
        return redirect(LISTAR_TIPOS_PUBLICACION_URL)

    return render(
        request,
        "admin/Publicaciones/editTipoPublicacion.html",
        {"tipo": tipo, "estados": cons.Estado.choices},
    )


def _build_inventario_json():
    from apps.inventory.models import Inventario
    inv_items = list(Inventario.objects.select_related("material__categoria", "punto_eca").all())
    return [{
        "matId": str(inv.material_id),
        "puntoId": str(inv.punto_eca_id),
        "stock": float(inv.stock_actual or 0),
        "cap": float(inv.capacidad_maxima or 0),
        "compra": float(inv.precio_compra or 0),
        "venta": float(inv.precio_venta or 0),
        "unidad": inv.unidad_medida,
    } for inv in inv_items]


def _build_puntos_json():
    from apps.ecas.models import PuntoECA
    puntos_qs = list(PuntoECA.objects.all().order_by("nombre"))
    return [{"id": str(p.id), "nombre": p.nombre, "localidad": p.localidad.nombre if p.localidad else "", "gestor": p.gestor_eca.get_full_name() if p.gestor_eca else "", "estado": p.estado} for p in puntos_qs]


def _compra_to_dict(c):
    inv = c.inventario
    return {
        "fecha": c.fecha_compra.strftime("%Y-%m-%d %H:%M") if c.fecha_compra else "",
        "tipo": "Compra", "mat": inv.material.nombre if inv and inv.material else "-",
        "punto": inv.punto_eca.nombre if inv and inv.punto_eca else "-",
        "kg": float(c.cantidad or 0), "unitario": float(c.precio_compra or 0),
        "total": float((c.cantidad or 0) * (c.precio_compra or 0)),
        "centro": inv.centro_acopio.nombre if inv and inv.centro_acopio else "-",
    }


def _venta_to_dict(v):
    inv = v.inventario
    return {
        "fecha": v.fecha_venta.strftime("%Y-%m-%d %H:%M") if v.fecha_venta else "",
        "tipo": "Venta", "mat": inv.material.nombre if inv and inv.material else "-",
        "punto": inv.punto_eca.nombre if inv and inv.punto_eca else "-",
        "kg": float(v.cantidad or 0), "unitario": float(v.precio_venta or 0),
        "total": float((v.cantidad or 0) * (v.precio_venta or 0)),
        "centro": v.centro_acopio.nombre if v.centro_acopio else "-",
    }


def _build_historial_json():
    from apps.operations.models import CompraInventario, VentaInventario
    compras = list(CompraInventario.objects.select_related("inventario__material", "inventario__punto_eca").order_by("-fecha_compra")[:200])
    ventas = list(VentaInventario.objects.select_related("inventario__material", "inventario__punto_eca").order_by("-fecha_venta")[:200])
    return [_compra_to_dict(c) for c in compras] + [_venta_to_dict(v) for v in ventas]


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def gestion_materiales(request):
    tab = request.GET.get('tab', 'materiales')

    all_materiales = list(Material.objects.select_related("categoria").all().order_by("nombre"))
    categorias_qs = CategoriaMaterial.objects.all().order_by("nombre")

    materiales_json = [material_to_dict(m) for m in all_materiales]
    categorias_json = [categoria_material_to_dict(c) for c in categorias_qs]

    return render(request, "admin/materiales_gestion.html", {
        "materiales_json": materiales_json,
        "categorias_json": categorias_json,
        "inventario_json": _build_inventario_json(),
        "puntos_json": _build_puntos_json(),
        "historial_json": _build_historial_json(),
        "tab": tab,
        "estados": cons.Estado.choices,
        "todas_categorias": CategoriaMaterial.objects.all().order_by("nombre"),
        "dist_categoria_mat": AdminDashboardService.obtener_distribucion_materiales(),
        "dist_clasificacion_mat": AdminDashboardService.obtener_distribucion_materiales_por_clasificacion(),
        "dist_estado_mat": AdminDashboardService.obtener_distribucion_materiales_por_estado(),
    })


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "POST"])
def crear_categoria_publicacion(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    context = {
        "tipos_publicacion": AdminCatalogService._tipos_publicacion_disponibles(),
        "form_data": {
            "nombre": "",
            "descripcion": "",
            "tipo": "",
            "estado": "ACTIVO",
        },
        "active_tab": "categorias_publicacion",
    }
    if request.method == "POST":
        resultado = AdminCatalogService.crear_categoria_publicacion(request.POST)
        if is_ajax:
            return JsonResponse(resultado)
        if resultado["ok"]:
            messages.success(request, resultado["message"])
            return redirect(LISTAR_CATEGORIAS_PUBLICACION_URL)
        messages.error(request, resultado["message"])
        context["form_data"] = {
            "nombre": request.POST.get("nombre", ""),
            "descripcion": request.POST.get("descripcion", ""),
            "tipo": request.POST.get("tipo", ""),
            "estado": request.POST.get("estado", "ACTIVO"),
        }

    return render(request, "admin/CategoriasPublicaciones/createCategoriaPublicacion.html", context)


# ─── Perfil del Administrador ───────────────────────────────────────────────

_CELULAR       = _re.compile(r"^3\d{9}$")
_PASSWORD_COMP = _re.compile(
    r"^(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[@$!%*?&_])[A-Za-z\d@$!%*?&_]{8,128}$"
)


@require_GET
@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_http_methods(["GET", "HEAD"])
def perfil_admin(request):
    localidades = Localidad.objects.all()
    tipos_documento = cons.TipoDocumento.choices
    return render(request, "admin/perfil_admin.html", {
        "localidades": localidades,
        "tipos_documento": tipos_documento,
    })


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_POST
def actualizar_datos_admin(request):

    is_ajax_request = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or "application/json" in request.headers.get("accept", "")
    )

    def _finish(ok, message, status_code=200):
        if is_ajax_request:
            return JsonResponse({"ok": ok, "message": message}, status=status_code)
        if ok:
            messages.success(request, message)
        else:
            messages.error(request, message)
        return redirect(ADMIN_PERFIL_URL)

    user = request.user
    nombres          = request.POST.get("nombres", "").strip()
    apellidos        = request.POST.get("apellidos", "").strip()
    celular          = request.POST.get("celular", "").strip()
    ciudad           = request.POST.get("ciudad", "").strip()
    localidad_id     = request.POST.get("localidad", "").strip()
    fecha_str        = request.POST.get("fechaNacimiento", "").strip()
    email            = request.POST.get("email", "").strip().lower()
    tipo_documento   = request.POST.get("tipoDocumento", "").strip()
    numero_documento = request.POST.get("numeroDocumento", "").strip()

    errores, fecha_nacimiento, localidad_inst = _validar_datos_perfil_admin(
        user, nombres, apellidos, celular, ciudad, fecha_str, localidad_id,
        email, tipo_documento, numero_documento,
    )

    if errores:
        return _finish(False, errores[0], status_code=400)

    try:
        user.nombres          = nombres
        user.apellidos        = apellidos
        user.celular          = celular if celular else None
        user.ciudad           = ciudad if ciudad else DEFAULT_CITY
        user.localidad        = localidad_inst
        user.fecha_nacimiento = fecha_nacimiento
        if email:
            user.email = email
        if tipo_documento:
            user.tipo_documento = tipo_documento
        if numero_documento:
            user.numero_documento = numero_documento
        user.save()
    except Exception:
        return _finish(False, "No se pudieron guardar los cambios.", status_code=500)

    return _finish(True, "Datos actualizados correctamente.")


@login_required(login_url="/login/")
@user_passes_test(es_administrador, login_url="/inicio/")
@require_POST
def cambiar_contrasena_admin(request):

    is_ajax_request = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or "application/json" in request.headers.get("accept", "")
    )

    def _finish(ok, message, status_code=200):
        if is_ajax_request:
            return JsonResponse({"ok": ok, "message": message}, status=status_code)
        if ok:
            messages.success(request, message)
        else:
            messages.error(request, message)
        return redirect(ADMIN_PERFIL_URL)

    user      = request.user
    actual    = request.POST.get("contrasenaActual", "")
    nueva     = request.POST.get("contrasenaNueva", "")
    confirmar = request.POST.get("confirmarContrasena", "")

    error = _validar_cambio_contrasena_admin(user, actual, nueva, confirmar)
    if error:
        return _finish(False, error, status_code=400)

    user.set_password(nueva)
    user.save()
    update_session_auth_hash(request, user)
    return _finish(True, "Contraseña actualizada correctamente.")
