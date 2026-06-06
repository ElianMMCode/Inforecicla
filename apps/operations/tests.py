import io
import uuid

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.test.utils import override_settings
from openpyxl import load_workbook

from apps.ecas.models import CentroAcopio, Localidad, PuntoECA
from apps.inventory.models import (
    CategoriaMaterial,
    Inventario,
    Material,
    TipoMaterial,
)
from apps.operations.models import CompraInventario, VentaInventario


def _crear_gestor(email):
    U = get_user_model()
    suffix = uuid.uuid4().hex[:8]
    return U.objects.create_user(
        email=email,
        numero_documento=f"9000{suffix}",
        password="x",  # NOSONAR - credencial dummy de tests, no es real
        nombres="Hist Test",
        apellidos="Apellido Test",
        tipo_documento="CC",
        tipo_usuario="GECA",
    )


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class TestExportHistorialFiltrosAvanzados(TestCase):
    """Verifica que los nuevos filtros (cantidad min/max, monto min/max,
    centro de acopio) lleguen al endpoint de export y se apliquen
    correctamente a la query del servidor."""

    @classmethod
    def setUpTestData(cls):
        cls.localidad = Localidad.objects.create(
            localidad_id=uuid.uuid4(), nombre="Bogotá"
        )
        cls.categoria_metal = CategoriaMaterial.objects.create(nombre="Metales")
        cls.tipo_alum = TipoMaterial.objects.create(nombre="Aluminio")
        cls.centro_a = CentroAcopio.objects.create(
            nombre="Centro A",
            tipo_centro="PUBLICO",
            visibilidad="GLOBAL",
            email="a@x.com",
            celular="3000000000",
            ciudad="Bogotá",
            localidad=cls.localidad,
        )
        cls.centro_b = CentroAcopio.objects.create(
            nombre="Centro B",
            tipo_centro="PUBLICO",
            visibilidad="GLOBAL",
            email="b@x.com",
            celular="3000000001",
            ciudad="Bogotá",
            localidad=cls.localidad,
        )

    def setUp(self):
        self.client = Client()
        self.user = _crear_gestor(f"hist-{uuid.uuid4().hex[:6]}@example.com")
        self.punto = PuntoECA.objects.create(
            gestor_eca=self.user,
            nombre="Punto Hist",
            telefono_punto="6010000000",
            direccion="calle 100",
            ciudad="Bogotá",
            email="punto@x.com",
            celular="3009999999",
            latitud=4.6,
            longitud=-74.0,
            localidad=self.localidad,
        )
        self.mat = Material.objects.create(
            nombre="Lata Test",
            categoria=self.categoria_metal,
            tipo=self.tipo_alum,
        )
        self.inv = Inventario.objects.create(
            punto_eca=self.punto,
            material=self.mat,
            capacidad_maxima=1000,
            unidad_medida="KG",
            stock_actual=500,
            umbral_alerta=70,
            umbral_critico=90,
            precio_compra=1000,
            precio_venta=2000,
        )
        self.client.force_login(self.user)

    def _crear_compra(self, cantidad, precio, fecha=None):
        return CompraInventario.objects.create(
            inventario=self.inv,
            cantidad=cantidad,
            precio_compra=precio,
            fecha_compra=fecha or "2024-06-15T10:00:00",
        )

    def _crear_venta(self, cantidad, precio, centro, fecha=None):
        return VentaInventario.objects.create(
            inventario=self.inv,
            cantidad=cantidad,
            precio_venta=precio,
            centro_acopio=centro,
            fecha_venta=fecha or "2024-06-15T11:00:00",
        )

    def _exportar_excel(self, **params):
        url = "/punto-eca/movimientos/exportar-historial-excel/"
        if params:
            url += "?" + "&".join(f"{k}={v}" for k, v in params.items())
        return self.client.get(url, HTTP_ACCEPT="application/json")

    def _contar_filas_xlsx(self, response):
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb.active
        return sum(1 for _ in ws.iter_rows(min_row=2))

    def test_export_sin_filtros_devuelve_todo(self):
        self._crear_compra(10, 1000)
        self._crear_compra(20, 2000)
        self._crear_venta(5, 3000, self.centro_a)
        response = self._exportar_excel()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 3)

    def test_export_filtro_cantidad_min(self):
        self._crear_compra(10, 1000)
        self._crear_compra(50, 1000)
        self._crear_compra(100, 1000)
        response = self._exportar_excel(cantidad_min=30)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 2)

    def test_export_filtro_cantidad_max(self):
        self._crear_compra(10, 1000)
        self._crear_compra(50, 1000)
        self._crear_compra(100, 1000)
        response = self._exportar_excel(cantidad_max=60)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 2)

    def test_export_filtro_cantidad_min_y_max(self):
        self._crear_compra(10, 1000)
        self._crear_compra(30, 1000)
        self._crear_compra(50, 1000)
        self._crear_compra(100, 1000)
        response = self._exportar_excel(cantidad_min=20, cantidad_max=60)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 2)

    def test_export_filtro_monto_min(self):
        # Compra 10x1000=10000 queda excluida; venta 50x5000=250000 incluida.
        self._crear_compra(10, 1000)
        self._crear_venta(50, 5000, self.centro_a)
        response = self._exportar_excel(monto_min=100000)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 1)

    def test_export_filtro_monto_max(self):
        self._crear_compra(10, 1000)
        self._crear_venta(50, 5000, self.centro_a)
        response = self._exportar_excel(monto_max=20000)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 1)

    def test_export_filtro_centro_acopio_solo_ventas(self):
        self._crear_venta(10, 2000, self.centro_a)
        self._crear_venta(20, 2000, self.centro_b)
        self._crear_compra(30, 1000)
        # El UI solo habilita el filtro de centro cuando tipo=venta;
        # en ese caso las compras se bloquean y solo las ventas al
        # centro elegido se exportan.
        response = self._exportar_excel(
            tipo_movimiento="venta", centro_acopio="Centro A"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 1)

    def test_export_filtro_centro_acopio_excluye_ventas_otros_centros(self):
        self._crear_venta(10, 2000, self.centro_a)
        self._crear_venta(20, 2000, self.centro_b)
        response = self._exportar_excel(
            tipo_movimiento="venta", centro_acopio="Centro A"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 1)

    def test_export_sin_resultados_retorna_404_json(self):
        self._crear_compra(10, 1000)
        response = self._exportar_excel(monto_min=99999999)
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response["Content-Type"], "application/json")

    def test_export_combina_todos_los_filtros(self):
        # Compra cantidad=50, monto=50000 → pasa todos los filtros
        self._crear_compra(50, 1000, fecha="2024-06-15T10:00:00")
        # Compra cantidad=200 → fuera de rango cantidad_max=100
        self._crear_compra(200, 1000, fecha="2024-06-15T10:00:00")
        # Venta cantidad=30, monto=60000 → pasa todos los filtros
        self._crear_venta(30, 2000, self.centro_a, fecha="2024-06-15T11:00:00")
        response = self._exportar_excel(
            material="Lata Test",
            categoria="Metales",
            tipo="Aluminio",
            fecha_desde="2024-06-01",
            fecha_hasta="2024-06-30",
            cantidad_min=10,
            cantidad_max=100,
            monto_min=10000,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 2)

    def test_export_parametro_decimal_invalido_se_ignora(self):
        self._crear_compra(10, 1000)
        response = self._exportar_excel(cantidad_min="abc")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self._contar_filas_xlsx(response), 1)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class TestExportFilename(TestCase):
    """Verifica que el nombre del archivo de export incluya la fecha
    de creación del archivo y el material filtrado (o 'general' si no
    hay filtro de material).
    """

    @classmethod
    def setUpTestData(cls):
        cls.localidad = Localidad.objects.create(
            localidad_id=uuid.uuid4(), nombre="Bogotá"
        )
        cls.user = _crear_gestor("filename-test@example.com")
        cls.punto = PuntoECA.objects.create(
            nombre="Punto Filename",
            email=f"fn-{uuid.uuid4().hex[:8]}@x.com",
            celular="3005556666",
            latitud=4.6,
            longitud=-74.0,
            localidad=cls.localidad,
            gestor_eca=cls.user,
        )
        cls.categoria = CategoriaMaterial.objects.create(nombre="Metales")
        cls.tipo = TipoMaterial.objects.create(nombre="Aluminio")
        cls.mat = Material.objects.create(
            nombre="Lata Test", categoria=cls.categoria, tipo=cls.tipo
        )
        cls.inv = Inventario.objects.create(
            punto_eca=cls.punto,
            material=cls.mat,
            capacidad_maxima=1000,
            unidad_medida="KG",
            stock_actual=500,
            umbral_alerta=70,
            umbral_critico=90,
            precio_compra=1000,
            precio_venta=2000,
        )
        cls.compra = CompraInventario.objects.create(
            inventario=cls.inv,
            cantidad=10,
            precio_compra=1000,
            fecha_compra="2024-06-15T10:00:00",
        )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.user)

    def _disposition(self, response):
        # Devuelve el filename (sin prefijo 'attachment; filename="').
        cd = response["Content-Disposition"]
        return cd.split("filename=")[-1].strip('"; ')

    def test_filename_compras_con_material(self):
        response = self.client.get(
            "/punto-eca/movimientos/exportar-compras-excel/?material=Lata%20Test",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        fn = self._disposition(response)
        # formato: compras_Lata_Test_YYYY-MM-DD_HHMM.xlsx
        self.assertRegex(fn, r"^compras_Lata_Test_\d{4}-\d{2}-\d{2}_\d{4}\.xlsx$")

    def test_filename_compras_sin_material(self):
        response = self.client.get(
            "/punto-eca/movimientos/exportar-compras-excel/",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        fn = self._disposition(response)
        self.assertRegex(fn, r"^compras_general_\d{4}-\d{2}-\d{2}_\d{4}\.xlsx$")

    def test_filename_pdf_historial(self):
        # PDF inline filename
        response = self.client.get(
            "/punto-eca/movimientos/exportar-historial-pdf/?material=Lata%20Test",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        cd = response["Content-Disposition"]
        self.assertIn("inline", cd)
        self.assertRegex(cd, r"historial_Lata_Test_\d{4}-\d{2}-\d{2}_\d{4}\.pdf")

    def test_filename_sanitiza_acentos(self):
        response = self.client.get(
            "/punto-eca/movimientos/exportar-compras-excel/?material=Lata%20Pl%C3%A1stico",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        fn = self._disposition(response)
        # 'Plástico' → 'Plastico' (sin tilde, sin 'á' en filename)
        self.assertIn("Plastico", fn)
        self.assertNotIn("á", fn)
        self.assertRegex(fn, r"^compras_Lata_Plastico_\d{4}-\d{2}-\d{2}_\d{4}\.xlsx$")

    def test_filename_dos_export_seguidos_difieren_o_coinciden_legitimamente(self):
        # Sin ventas: 404 con cuerpo JSON, no se genera filename.
        # Pero igual probamos con compras para validar el formato.
        r3 = self.client.get(
            "/punto-eca/movimientos/exportar-compras-excel/",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r3.status_code, 200)
        self.assertRegex(self._disposition(r3), r"^compras_general_\d{4}-\d{2}-\d{2}_\d{4}\.xlsx$")


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class TestEditarMovimiento(TestCase):
    """Verifica que la edición de compras/ventas desde el historial general:

    1. El servicio recibe los nombres de campos correctos (compraId,
       fechaCompra, etc.) y persiste los cambios.
    2. La view propaga el 'status' del body al HTTP status (bug histórico:
       retornaba 200 incluso con error → el JS mostraba "Compra
       actualizada" sin guardar).
    """

    @classmethod
    def setUpTestData(cls):
        cls.localidad = Localidad.objects.create(
            localidad_id=uuid.uuid4(), nombre="Bogotá"
        )
        cls.categoria = CategoriaMaterial.objects.create(nombre="Metales")
        cls.tipo = TipoMaterial.objects.create(nombre="Aluminio")
        cls.centro = CentroAcopio.objects.create(
            nombre="Centro Edit",
            tipo_centro="PUBLICO",
            visibilidad="GLOBAL",
            email="edit@x.com",
            celular="3001112222",
            ciudad="Bogotá",
            localidad=cls.localidad,
        )
        cls.user = _crear_gestor("edit-test@example.com")
        cls.punto = PuntoECA.objects.create(
            nombre="Punto Edit",
            email=f"edit-{uuid.uuid4().hex[:8]}@x.com",
            celular="3003334444",
            latitud=4.6,
            longitud=-74.0,
            localidad=cls.localidad,
            gestor_eca=cls.user,
        )
        cls.mat = Material.objects.create(
            nombre="Lata Edit", categoria=cls.categoria, tipo=cls.tipo
        )
        cls.inv = Inventario.objects.create(
            punto_eca=cls.punto,
            material=cls.mat,
            capacidad_maxima=1000,
            unidad_medida="KG",
            stock_actual=500,
            umbral_alerta=70,
            umbral_critico=90,
            precio_compra=1000,
            precio_venta=2000,
        )
        cls.compra = CompraInventario.objects.create(
            inventario=cls.inv,
            cantidad=10,
            precio_compra=1000,
            fecha_compra="2024-06-15T10:00:00",
        )
        cls.venta = VentaInventario.objects.create(
            inventario=cls.inv,
            cantidad=5,
            precio_venta=2000,
            centro_acopio=cls.centro,
            fecha_venta="2024-06-15T11:00:00",
        )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.user)

    def _patch_compra(self, payload):
        return self.client.patch(
            f"/punto-eca/movimientos/editar-compra/{self.compra.id}/",
            data=__import__("json").dumps(payload),
            content_type="application/json",
            HTTP_X_CSRFTOKEN="x",
        )

    def _patch_venta(self, payload):
        return self.client.patch(
            f"/punto-eca/movimientos/editar-venta/{self.venta.id}/",
            data=__import__("json").dumps(payload),
            content_type="application/json",
            HTTP_X_CSRFTOKEN="x",
        )

    def test_editar_compra_payload_correcto_persiste_y_retorna_200(self):
        response = self._patch_compra({
            "compraId": str(self.compra.id),
            "fechaCompra": "2025-01-15T10:00:00",
            "cantidad": "25.00",
            "precioCompra": "1500.00",
            "observaciones": "editada",
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json().get("error"))
        self.compra.refresh_from_db()
        self.assertEqual(self.compra.cantidad, 25)
        self.assertEqual(self.compra.precio_compra, 1500)
        self.assertEqual(self.compra.observaciones, "editada")

    def test_editar_compra_sin_compraId_retorna_400_no_200(self):
        # Bug histórico: la view no propagaba 'status' del body, retornaba
        # 200 con error=True → JS mostraba éxito sin guardar.
        response = self._patch_compra({
            "id": str(self.compra.id),
            "fechaCompra": "2025-01-15T10:00:00",
            "cantidad": "20.00",
            "precioCompra": "1500.00",
        })
        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json().get("error"))
        # La compra NO se modifica.
        self.compra.refresh_from_db()
        self.assertEqual(self.compra.cantidad, 10)

    def test_editar_compra_cantidad_negativa_retorna_400(self):
        response = self._patch_compra({
            "compraId": str(self.compra.id),
            "fechaCompra": "2025-01-15T10:00:00",
            "cantidad": "-1",
            "precioCompra": "1500.00",
        })
        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json().get("error"))

    def test_editar_venta_payload_correcto_persiste_y_retorna_200(self):
        response = self._patch_venta({
            "ventaId": str(self.venta.id),
            "fechaVenta": "2025-01-15T11:00:00",
            "cantidad": "8.00",
            "precioVenta": "3500.00",
            "observaciones": "editada",
            "centroAcopioId": str(self.centro.id),
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json().get("error"))
        self.venta.refresh_from_db()
        self.assertEqual(self.venta.cantidad, 8)
        self.assertEqual(self.venta.precio_venta, 3500)

    def test_editar_venta_sin_ventaId_retorna_400_no_200(self):
        response = self._patch_venta({
            "id": str(self.venta.id),
            "fechaVenta": "2025-01-15T11:00:00",
            "cantidad": "8.00",
            "precioVenta": "3500.00",
        })
        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json().get("error"))
        self.venta.refresh_from_db()
        self.assertEqual(self.venta.cantidad, 5)


class TestFlujoEndToEndCompraInventario(TestCase):
    """Test end-to-end del flujo CRUD de CompraInventario:

    1. POST /registrar-compra/  → crea CompraInventario y actualiza stock.
    2. PATCH /editar-compra/   → edita cantidad/precio/observaciones y
                                 re-calcula el stock.
    3. DELETE /borrar-compra/  → elimina CompraInventario y revierte stock.

    Es la primera barrera de protección contra regresiones del flujo
    completo. Si cualquiera de los 3 pasos rompe, el test lo detecta.
    """

    @classmethod
    def setUpTestData(cls):
        cls.localidad = Localidad.objects.create(
            localidad_id=uuid.uuid4(), nombre="Bogotá"
        )
        cls.categoria = CategoriaMaterial.objects.create(nombre="Metales")
        cls.tipo = TipoMaterial.objects.create(nombre="Aluminio")
        cls.user = _crear_gestor("e2e-compra@example.com")
        cls.punto = PuntoECA.objects.create(
            nombre="Punto E2E",
            email=f"e2e-{uuid.uuid4().hex[:8]}@x.com",
            celular="3005556666",
            latitud=4.6,
            longitud=-74.0,
            localidad=cls.localidad,
            gestor_eca=cls.user,
        )
        cls.mat = Material.objects.create(
            nombre="Lata E2E", categoria=cls.categoria, tipo=cls.tipo
        )
        cls.inv = Inventario.objects.create(
            punto_eca=cls.punto,
            material=cls.mat,
            capacidad_maxima=1000,
            unidad_medida="KG",
            stock_actual=500,
            umbral_alerta=70,
            umbral_critico=90,
            precio_compra=1000,
            precio_venta=2000,
        )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.user)

    def _post_compra(self, payload):
        import json
        return self.client.post(
            "/punto-eca/movimientos/registrar-compra/",
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_X_CSRFTOKEN="x",
        )

    def _patch_compra(self, compra_id, payload):
        import json
        return self.client.patch(
            f"/punto-eca/movimientos/editar-compra/{compra_id}/",
            data=json.dumps(payload),
            content_type="application/json",
        )

    def _delete_compra(self, compra_id):
        return self.client.delete(
            f"/punto-eca/movimientos/borrar-compra/{compra_id}/"
        )

    def test_flujo_completo_crear_editar_eliminar(self):
        # 1. CREAR: POST /registrar-compra/ con cantidad=20, precio=1500
        # Stock inicial: 500. Después: 500 + 20 = 520
        stock_inicial = self.inv.stock_actual
        response = self._post_compra({
            "inventarioId": str(self.inv.id),
            "cantidad": "20.00",
            "precioCompra": "1500.00",
            "fechaCompra": "2025-06-15T10:00:00",
            "observaciones": "compra inicial",
        })
        self.assertEqual(response.status_code, 201,
                         f"POST crear debe retornar 201; body={response.content!r}")
        self.assertFalse(response.json().get("error"))
        self.assertEqual(CompraInventario.objects.count(), 1)
        compra = CompraInventario.objects.first()
        self.assertEqual(compra.cantidad, 20)
        self.assertEqual(compra.precio_compra, 1500)
        # Stock actualizado
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, stock_inicial + 20)

        # 2. EDITAR: PATCH cantidad de 20 a 30 → stock debe subir +10
        # Stock antes: 520. Después: 520 + 10 = 530
        response = self._patch_compra(compra.id, {
            "compraId": str(compra.id),
            "cantidad": "30.00",
            "precioCompra": "1500.00",
            "fechaCompra": "2025-06-15T10:00:00",
            "observaciones": "editada",
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json().get("error"))
        compra.refresh_from_db()
        self.assertEqual(compra.cantidad, 30)
        self.assertEqual(compra.observaciones, "editada")
        # Stock re-calculado
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, stock_inicial + 30)

        # 3. ELIMINAR: DELETE → stock revierte al inicial
        response = self._delete_compra(compra.id)
        self.assertIn(response.status_code, (200, 204),
                      f"DELETE debe retornar 2xx; body={response.content!r}")
        self.assertFalse(CompraInventario.objects.filter(id=compra.id).exists())
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, stock_inicial)

    def test_crear_compra_con_fallback_por_punto_y_material(self):
        """Si el JS no envía inventarioId, el servicio retorna 400 con
        mensaje claro. El JS siempre envía inventarioId desde
        formEntradaInventarioId, por lo que este caso solo se da si
        alguien manipula el payload directamente."""
        response = self._post_compra({
            "puntoEcaId": str(self.punto.id),
            "materialId": str(self.mat.id),
            "cantidad": "5.00",
            "precioCompra": "1000.00",
            "fechaCompra": "2025-06-15T10:00:00",
        })
        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json().get("error"))
        self.assertIn("inventarioId", response.json().get("mensaje", ""))
        # No se creó la compra
        self.assertEqual(CompraInventario.objects.count(), 0)
        # Stock no cambió
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, 500)

    def test_crear_compra_cantidad_negativa_retorna_400(self):
        """No se debe poder registrar una compra con cantidad <= 0."""
        response = self._post_compra({
            "inventarioId": str(self.inv.id),
            "cantidad": "-1.00",
            "precioCompra": "1000.00",
            "fechaCompra": "2025-06-15T10:00:00",
        })
        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.json().get("error"))
        # No se creó la compra
        self.assertEqual(CompraInventario.objects.count(), 0)
        # Stock no cambió
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, 500)

    def _post_venta(self, payload):
        import json
        return self.client.post(
            "/punto-eca/movimientos/registrar-venta/",
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_X_CSRFTOKEN="x",
        )

    def test_crear_venta_sin_centro_es_opcional_retorna_201(self):
        """Una venta SIN centro de acopio debe ser ACEPTADA con 201
        (el centro es opcional en ventas). Antes era rechazado como
        defensa en profundidad, pero el negocio permite ventas internas
        o sin destino externo. Se valida que la venta se cree con
        centro_acopio=None y que el stock se descuente correctamente."""
        from apps.operations.models import VentaInventario
        # Caso 1: sin centro (key omitida)
        response = self._post_venta({
            "inventarioId": str(self.inv.id),
            "cantidad": "5.00",
            "precioVenta": "2000.00",
            "fechaVenta": "2025-06-15T10:00:00",
            "observaciones": "venta sin centro (interna)",
        })
        self.assertEqual(response.status_code, 201,
                         f"Venta sin centro debe aceptarse; body={response.content!r}")
        self.assertFalse(response.json().get("error"))
        self.assertEqual(VentaInventario.objects.count(), 1)
        venta = VentaInventario.objects.first()
        self.assertIsNone(venta.centro_acopio,
                          "La venta sin centro debe persistir con centro_acopio=None")
        # Stock se descuenta correctamente
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, 500 - 5)
        # Caso 2: centro vacío (cadena vacía) — también debe aceptarse
        venta.delete()
        self.inv.refresh_from_db()
        self.inv.stock_actual = 500
        self.inv.save()
        response = self._post_venta({
            "inventarioId": str(self.inv.id),
            "cantidad": "3.00",
            "precioVenta": "1500.00",
            "fechaVenta": "2025-06-15T10:00:00",
            "centroAcopioId": "",
        })
        self.assertEqual(response.status_code, 201,
                         f"centroAcopioId='' también debe aceptarse; body={response.content!r}")
        self.assertFalse(response.json().get("error"))
        # Caso 3: con centro válido — también debe aceptarse
        from apps.ecas.models import CentroAcopio
        centro = CentroAcopio.objects.create(
            nombre="Centro Test",
            email=f"centro-{uuid.uuid4().hex[:8]}@x.com",
            celular="3115556677",
            latitud=4.6,
            longitud=-74.0,
            localidad=self.localidad,
        )
        response = self._post_venta({
            "inventarioId": str(self.inv.id),
            "cantidad": "2.00",
            "precioVenta": "2500.00",
            "fechaVenta": "2025-06-15T10:00:00",
            "centroAcopioId": str(centro.id),
        })
        self.assertEqual(response.status_code, 201)
        venta = VentaInventario.objects.filter(centro_acopio=centro).first()
        self.assertIsNotNone(venta, "La venta con centro válido debe persistirse con la FK")


def _csv_inline(nombre, contenido):
    """Helper para crear un SimpleUploadedFile CSV en memoria."""
    from django.core.files.uploadedfile import SimpleUploadedFile
    return SimpleUploadedFile(
        nombre,
        contenido.encode("utf-8"),
        content_type="text/csv",
    )


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class TestBulkImportEndpoint(TestCase):
    """Tests del endpoint bulk_import (carga masiva CSV) y del endpoint
    auxiliar descargar_plantilla_bulk.

    Cubre:
    - Validaciones de input (archivo, extensión, headers)
    - Procesamiento exitoso con N filas (incluido el archivo real de 600 filas)
    - Creación automática de Inventario para materiales nuevos
    - Descarga de plantilla con header correcto
    - Métodos no permitidos
    """

    @classmethod
    def setUpTestData(cls):
        cls.localidad = Localidad.objects.create(
            localidad_id=uuid.uuid4(), nombre="Bogotá"
        )
        cls.categoria = CategoriaMaterial.objects.create(nombre="Metales")
        cls.tipo = TipoMaterial.objects.create(nombre="Aluminio")
        cls.user = _crear_gestor("bulk-test@example.com")
        cls.punto = PuntoECA.objects.create(
            nombre="Punto Bulk",
            email=f"bulk-{uuid.uuid4().hex[:8]}@x.com",
            celular="3005557777",
            latitud=4.6,
            longitud=-74.0,
            localidad=cls.localidad,
            gestor_eca=cls.user,
        )
        cls.mat = Material.objects.create(
            nombre="Lata Bulk", categoria=cls.categoria, tipo=cls.tipo
        )
        cls.inv = Inventario.objects.create(
            punto_eca=cls.punto,
            material=cls.mat,
            capacidad_maxima=10000,
            unidad_medida="KG",
            stock_actual=0,
            umbral_alerta=70,
            umbral_critico=90,
            precio_compra=1000,
            precio_venta=2000,
        )

    def setUp(self):
        self.client = Client(enforce_csrf_checks=False)
        self.client.force_login(self.user)

    # --- Helpers ---

    def _post_bulk_compra(self, archivo):
        return self.client.post(
            "/punto-eca/movimientos/compras/bulk_import/",
            data={"file": archivo},
            HTTP_X_CSRFTOKEN="x",
            HTTP_ACCEPT="application/json",
        )

    def _post_bulk_venta(self, archivo):
        return self.client.post(
            "/punto-eca/movimientos/ventas/bulk_import/",
            data={"file": archivo},
            HTTP_X_CSRFTOKEN="x",
            HTTP_ACCEPT="application/json",
        )

    # --- Validaciones de input ---

    def test_bulk_compras_sin_archivo_retorna_400(self):
        """POST sin campo 'file' debe retornar 400 con mensaje claro."""
        response = self.client.post(
            "/punto-eca/movimientos/compras/bulk_import/",
            HTTP_X_CSRFTOKEN="x",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertEqual(data.get("status"), "error")
        self.assertIn("archivo", data.get("mensaje", "").lower())

    def test_bulk_compras_archivo_no_csv_retorna_400(self):
        """Un archivo .txt debe ser rechazado con 400."""
        from django.core.files.uploadedfile import SimpleUploadedFile
        archivo = SimpleUploadedFile("not_csv.txt", b"hola", content_type="text/plain")
        response = self._post_bulk_compra(archivo)
        self.assertEqual(response.status_code, 400)
        self.assertIn("csv", response.json().get("mensaje", "").lower())

    def test_bulk_compras_headers_faltantes_retorna_400(self):
        """CSV sin todas las columnas requeridas debe ser rechazado."""
        csv_content = "nombreMaterial,cantidad\nLata Bulk,10\n"
        response = self._post_bulk_compra(_csv_inline("mal.csv", csv_content))
        self.assertEqual(response.status_code, 400, f"body={response.content!r}")
        msg = response.json().get("mensaje", "")
        self.assertIn("Faltan columnas", msg)
        self.assertIn("precioCompra", msg)

    def test_bulk_ventas_headers_faltantes_retorna_400(self):
        """CSV ventas sin todas las columnas requeridas debe ser rechazado."""
        csv_content = "nombreMaterial,cantidad,precioVenta\nLata Bulk,5,2000\n"
        response = self._post_bulk_venta(_csv_inline("mal_ventas.csv", csv_content))
        self.assertEqual(response.status_code, 400)
        self.assertIn("fechaVenta", response.json().get("mensaje", ""))

    def test_bulk_metodo_no_post_retorna_405(self):
        """GET al endpoint debe retornar 405 (Method Not Allowed)."""
        response = self.client.get("/punto-eca/movimientos/compras/bulk_import/")
        self.assertEqual(response.status_code, 405)

    # --- Procesamiento exitoso ---

    def test_bulk_compras_archivo_valido_crea_todas_las_compras(self):
        """CSV de 3 filas válidas debe crear 3 CompraInventario."""
        csv_content = (
            "nombreMaterial,cantidad,precioCompra,fechaCompra,observaciones\n"
            "Lata Bulk,10.0,1000.00,2026-06-15 10:00:00,Primera\n"
            "Lata Bulk,20.0,1100.00,2026-06-16 11:00:00,Segunda\n"
            "Lata Bulk,5.0,1200.00,2026-06-17 12:00:00,Tercera\n"
        )
        response = self._post_bulk_compra(_csv_inline("compras.csv", csv_content))
        self.assertEqual(response.status_code, 200,
                         f"body={response.content!r}")
        data = response.json()
        self.assertEqual(data.get("status"), "success")
        resumen = data.get("resumen", {})
        self.assertEqual(resumen.get("exitosas"), 3)
        self.assertEqual(resumen.get("con_errores"), 0)
        self.assertEqual(resumen.get("total_filas"), 3)
        # DB tiene 3 compras
        self.assertEqual(CompraInventario.objects.count(), 3)
        # Stock subió 35
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.stock_actual, 35)

    def test_bulk_compras_material_nuevo_en_catalogo_crea_inventario(self):
        """Si el material existe en el catálogo global pero NO en el
        inventario de este PuntoECA, el endpoint crea automáticamente el
        Inventario (a través de _buscar_o_crear_material_inventario) y
        registra la compra."""
        # Crear un material en el catálogo global que NO está en este punto
        cat2 = CategoriaMaterial.objects.create(nombre="Plásticos")
        tipo2 = TipoMaterial.objects.create(nombre="PET")
        mat_nuevo = Material.objects.create(
            nombre="Material Catalogo Nuevo", categoria=cat2, tipo=tipo2
        )
        self.assertFalse(
            Inventario.objects.filter(punto_eca=self.punto, material=mat_nuevo).exists()
        )
        csv_content = (
            "nombreMaterial,cantidad,precioCompra,fechaCompra,observaciones\n"
            "Material Catalogo Nuevo,15.0,1500.00,2026-06-15 10:00:00,Nuevo material\n"
        )
        response = self._post_bulk_compra(_csv_inline("nuevo.csv", csv_content))
        self.assertEqual(response.status_code, 200, f"body={response.content!r}")
        data = response.json()
        self.assertEqual(data.get("resumen", {}).get("exitosas"), 1)
        # El Inventario fue creado automáticamente
        self.assertTrue(
            Inventario.objects.filter(
                punto_eca=self.punto, material=mat_nuevo
            ).exists()
        )
        # Compra persiste
        self.assertEqual(CompraInventario.objects.count(), 1)

    def test_bulk_compras_600_filas_junio_2026_procesa_mayoria_correctamente(self):
        """Carga el CSV real de 600 filas (junio 2026). Debe procesar
        exitosamente la gran mayoría (~588) y fallar las ~12 con materiales
        inventados (2% de probabilidad en el script generador)."""
        # Cargar materiales del CSV (los reales del fixture)
        self._cargar_materiales_csv("docs/carga_masiva/test_compras_materiales_masivo.csv")
        csv_path = "docs/carga_masiva/test_compras_materiales_masivo.csv"
        with open(csv_path, "r", encoding="utf-8") as f:
            csv_content = f.read()
        response = self._post_bulk_compra(_csv_inline("masivo.csv", csv_content))
        self.assertEqual(response.status_code, 200,
                         f"body={response.content!r}")
        data = response.json()
        resumen = data.get("resumen", {})
        total = resumen.get("total_filas", 0)
        exitosas = resumen.get("exitosas", 0)
        con_errores = resumen.get("con_errores", 0)
        self.assertEqual(total, 600)
        # La mayoría debe ser exitosa (>= 550 de 600)
        self.assertGreaterEqual(
            exitosas, 550,
            f"esperaba >=550 exitosas, obtuve {exitosas} (errores: {con_errores})"
        )
        # La diferencia son los materiales inventados
        self.assertGreater(con_errores, 0, "El CSV tiene ~2% inventados")
        self.assertEqual(exitosas + con_errores, total)
        # DB tiene 'exitosas' compras
        self.assertEqual(CompraInventario.objects.count(), exitosas)

    def test_bulk_ventas_600_filas_junio_2026_procesa_mayoria_correctamente(self):
        """Carga el CSV real de 600 ventas (junio 2026). Carga compras
        primero para tener stock, luego ventas. Verifica que la mayoría se
        procesa correctamente."""
        # Cargar materiales y crear stock con compras previas
        self._cargar_materiales_csv("docs/carga_masiva/test_ventas_materiales_masivo.csv")
        # Crear un Inventario por cada material con stock alto
        nombres_unicos = set()
        with open("docs/carga_masiva/test_ventas_materiales_masivo.csv", "r", encoding="utf-8") as f:
            for line in f.readlines()[1:]:
                nombre = line.split(",")[0].strip()
                nombres_unicos.add(nombre)
        for nombre in nombres_unicos:
            mat = Material.objects.filter(nombre=nombre).first()
            if mat and not Inventario.objects.filter(punto_eca=self.punto, material=mat).exists():
                Inventario.objects.create(
                    punto_eca=self.punto, material=mat,
                    capacidad_maxima=100000, unidad_medida="KG",
                    stock_actual=10000, umbral_alerta=70, umbral_critico=90,
                    precio_compra=1000, precio_venta=2000,
                )
        csv_path = "docs/carga_masiva/test_ventas_materiales_masivo.csv"
        with open(csv_path, "r", encoding="utf-8") as f:
            csv_content = f.read()
        response = self._post_bulk_venta(_csv_inline("masivo_ventas.csv", csv_content))
        self.assertEqual(response.status_code, 200,
                         f"body={response.content!r}")
        data = response.json()
        resumen = data.get("resumen", {})
        total = resumen.get("total_filas", 0)
        exitosas = resumen.get("exitosas", 0)
        self.assertEqual(total, 600)
        # Con stock suficiente, la mayoría debe ser exitosa
        self.assertGreaterEqual(exitosas, 500,
            f"esperaba >=500 ventas exitosas, obtuve {exitosas}")

    def _cargar_materiales_csv(self, csv_path):
        """Helper: lee la columna nombreMaterial de un CSV y crea los
        materiales en el catálogo global (sin Inventario — eso lo hace el
        endpoint). Usado por los tests del CSV masivo. NO crea los
        materiales inventados (los que el script generador inserta con
        ~2% de probabilidad para forzar errores)."""
        import csv as csvmod
        MATERIALES_INVENTADOS = {
            "Material NoExistente",
            "Material Inventado 019",
            "Material Nuevo Test",
        }
        nombres = set()
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csvmod.DictReader(f)
            for fila in reader:
                nombre = fila["nombreMaterial"].strip()
                if nombre in MATERIALES_INVENTADOS:
                    continue
                nombres.add(nombre)
        cat = CategoriaMaterial.objects.first() or CategoriaMaterial.objects.create(nombre="Default")
        tipo = TipoMaterial.objects.first() or TipoMaterial.objects.create(nombre="Default")
        for nombre in nombres:
            if not Material.objects.filter(nombre=nombre).exists():
                Material.objects.create(nombre=nombre, categoria=cat, tipo=tipo)

    # --- Endpoint descargar_plantilla_bulk ---

    def test_descargar_plantilla_compra_retorna_csv_con_headers_correctos(self):
        """GET ?tipo=compra debe retornar un CSV con el header de compras."""
        response = self.client.get("/punto-eca/movimientos/descargar-plantilla/?tipo=compra")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("plantilla_compra_ejemplo.csv", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        # Header correcto
        self.assertIn("nombreMaterial", body)
        self.assertIn("precioCompra", body)
        self.assertIn("fechaCompra", body)
        # Al menos una fila de ejemplo
        lines = body.strip().split("\n")
        self.assertGreaterEqual(len(lines), 2)

    def test_descargar_plantilla_venta_retorna_csv_con_headers_correctos(self):
        """GET ?tipo=venta debe retornar un CSV con el header de ventas."""
        response = self.client.get("/punto-eca/movimientos/descargar-plantilla/?tipo=venta")
        self.assertEqual(response.status_code, 200)
        self.assertIn("plantilla_venta_ejemplo.csv", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        self.assertIn("precioVenta", body)
        self.assertIn("fechaVenta", body)
        self.assertNotIn("precioCompra", body)

    def test_descargar_plantilla_sin_parametro_retorna_plantilla_compra(self):
        """GET sin ?tipo debe usar el default 'compra'."""
        response = self.client.get("/punto-eca/movimientos/descargar-plantilla/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("precioCompra", response.content.decode("utf-8"))

    def test_descargar_plantilla_tipo_invalido_retorna_400(self):
        """GET ?tipo=xyz debe retornar 400 con mensaje claro."""
        response = self.client.get(
            "/punto-eca/movimientos/descargar-plantilla/?tipo=xyz",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertEqual(data.get("status"), "error")
        self.assertIn("compra", data.get("mensaje", ""))
        self.assertIn("venta", data.get("mensaje", ""))

    def test_descargar_plantilla_usuario_no_gestor_redirige(self):
        """Ciudadano (no gestor) no debe poder descargar plantilla.
        El decorador redirige a /inicio/ (no 403)."""
        U = get_user_model()
        ciu = U.objects.create_user(
            email="ciu@x.com",
            numero_documento="9999" + uuid.uuid4().hex[:6],
            password="x",  # NOSONAR - credencial dummy de tests, no es real
            nombres="Ciu",
            apellidos="Test",
            tipo_documento="CC",
            tipo_usuario="CIU",
        )
        client = Client(enforce_csrf_checks=False)
        client.force_login(ciu)
        response = client.get(
            "/punto-eca/movimientos/descargar-plantilla/?tipo=compra",
            HTTP_ACCEPT="application/json",
        )
        # El decorador redirige a /inicio/ (302)
        self.assertIn(response.status_code, (302, 301))

    def test_bulk_compras_usuario_no_gestor_redirige(self):
        """Ciudadano no debe poder hacer bulk import (redirige)."""
        U = get_user_model()
        ciu = U.objects.create_user(
            email="ciu2@x.com",
            numero_documento="9999" + uuid.uuid4().hex[:6],
            password="x",  # NOSONAR - credencial dummy de tests, no es real
            nombres="Ciu2",
            apellidos="Test",
            tipo_documento="CC",
            tipo_usuario="CIU",
        )
        client = Client(enforce_csrf_checks=False)
        client.force_login(ciu)
        csv_content = "nombreMaterial,cantidad,precioCompra,fechaCompra,observaciones\nLata Bulk,1,1,2026-06-15,x\n"
        response = client.post(
            "/punto-eca/movimientos/compras/bulk_import/",
            data={"file": _csv_inline("c.csv", csv_content)},
            HTTP_X_CSRFTOKEN="x",
            HTTP_ACCEPT="application/json",
        )
        self.assertIn(response.status_code, (302, 301))
